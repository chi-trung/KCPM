"""
generate_unittest_excel.py — v2 (đúng format UnitestCuaBao của thầy)

Format: mỗi Function sheet giống hệt ảnh chụp:
  Row2: Function Code | Function Name
  Row3: Created By   | Executed By
  Row4: Lines of code | Lack of test cases
  Row5: Test requirement
  Row6: Passed | Failed | Untested | N/A/B cols | Total Test Cases
  Row7: (count values)
  Row8: empty
  Row9: Condition | Precondition | (blank) | UTC01 | UTC02 | ...
  Row10+: condition rows with "O" marks
  ...
  Confirm section: Return codes, Exception, Log message
  Result section: Type(N/A/B), Passed/Failed, Executed Date, Defect ID

Sheet1: bảng tổng hợp tất cả test cases từ tất cả Function sheet
Coverage: bảng tổng hợp black-box coverage và white-box line/branch coverage
"""

import json
import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import date

TODAY = date.today().strftime("%d/%m/%Y")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_COVERAGE_SUMMARY = os.path.join(
    BASE_DIR, "Waste-Recycling-Platform", "frontend", "coverage", "coverage-summary.json"
)
BACKEND_COVERAGE_JSON = os.path.join(
    BASE_DIR, "Waste-Recycling-Platform", "backend", "coverage", "backend", "coverage.json"
)

# ─── Màu sắc ────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header section background
MED_BLUE    = "2E75B6"   # subheader / condition header
LIGHT_BLUE  = "BDD7EE"   # alternate row
LIGHT_GREEN = "E2EFDA"   # passed
LIGHT_RED   = "FFE0E0"   # failed
YELLOW      = "FFF2CC"   # UTCID header
GRAY_BG     = "F2F2F2"   # metadata rows
WHITE       = "FFFFFF"

# ─── Font chuẩn (Calibri hỗ trợ tiếng Việt tốt) ────────────────────
FONT_NAME = "Calibri"

# ─── Helper styles ─────────────────────────────────────────────────

def _side(color="000000"):
    return Side(style="thin", color=color)

THIN = Border(
    left=_side(), right=_side(), top=_side(), bottom=_side()
)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font_white_bold(sz=11):
    return Font(name=FONT_NAME, bold=True, color="FFFFFF", size=sz)

def font_bold(sz=11, color="000000"):
    return Font(name=FONT_NAME, bold=True, color=color, size=sz)

def font_normal(sz=11):
    return Font(name=FONT_NAME, size=sz)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN

def merge_set(ws, cell_range, value=None, fill=None, font=None, align=None):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    if value is not None:
        top_left.value = value
    if fill:
        top_left.fill = fill
    if font:
        top_left.font = font
    if align:
        top_left.alignment = align
    return top_left


TECHNIQUE_MAP = {
    "KIEM-4-F01":    "Black-box: Equivalence Partitioning, Error Guessing",
    "KIEM-5-F02":    "Black-box + White-box: BVA, Equivalence Partitioning, CFG/Path Coverage",
    "KIEM-8-F03":    "Black-box: State Transition Diagram",
    "KIEM-10-F04":   "Black-box: State Transition Diagram",
    "KIEM-16-F05":   "Black-box: Decision Table",
    "KIEM-13-F06":   "Black-box + White-box: Equivalence Partitioning, Error Guessing, CFG/Path Coverage",
    "KIEM-19-F07":   "Black-box: State Transition, Integration Test",
    "KIEM-20-F08":   "Black-box: BVA, Error Guessing",
    "KIEM-12-F09":   "Black-box: Equivalence Partitioning",
    "KIEM-FE-F10":   "Black-box: End-to-End Testing (CodeceptJS)",
    "KIEM-BVA-F11":  "Black-box: BVA Standard + Robustness",
    "KIEM-7-F12":    "Black-box: Decision Table Testing",
    "KIEM-5-F13":    "Black-box: State Transition Testing",
    "KIEM-FE-F14":   "White-box/Component: Vitest RTL + line/branch coverage",
    "KIEM-FE-F15":   "White-box/Component: Vitest RTL + line/branch coverage",
}

ROLE_MAP = {
    "KIEM-4-F01":    "Citizen / Enterprise",
    "KIEM-5-F02":    "Citizen",
    "KIEM-8-F03":    "Enterprise / Admin",
    "KIEM-10-F04":   "Collector",
    "KIEM-16-F05":   "Enterprise",
    "KIEM-13-F06":   "Citizen / Admin",
    "KIEM-19-F07":   "Citizen",
    "KIEM-20-F08":   "Collector",
    "KIEM-12-F09":   "Admin / Public",
    "KIEM-FE-F10":   "Citizen / Enterprise / Collector",
    "KIEM-BVA-F11":  "Citizen (upload ảnh bằng chứng)",
    "KIEM-7-F12":    "Citizen / Admin",
    "KIEM-5-F13":    "Citizen / Enterprise / Admin",
    "KIEM-FE-F14":   "Citizen / Enterprise / Admin (FE)",
    "KIEM-FE-F15":   "Citizen / Enterprise / Admin (FE)",
}

WHITEBOX_COVERAGE_MAP = {
    "KIEM-5-F02": {
        "source": "WastePlatform.Tests.Whitebox.CreateReportWhiteboxTests",
        "line": (10, 10),
        "branch": (10, 10),
        "statement": (10, 10),
        "condition": (8, 8),
        "path": (6, 6),
        "note": "CFG nodes=14, edges=16, V(G)=6; bao phủ đủ 6 independent paths.",
    },
    "KIEM-13-F06": {
        "source": "WastePlatform.Tests.Whitebox.EnterpriseRespondWhiteboxTests",
        "line": (10, 10),
        "branch": (10, 10),
        "statement": (10, 10),
        "condition": (4, 4),
        "path": (6, 6),
        "note": "CFG nodes=12, edges=12, V(G)=6; full truth table cho điều kiện trạng thái.",
    },
}

FRONTEND_COVERAGE_GROUPS = {
    "KIEM-FE-F14": [
        "src\\components\\shared\\StatCard.tsx",
        "src\\components\\shared\\ReportCard.tsx",
        "src\\components\\shared\\TaskCard.tsx",
        "src\\components\\shared\\ConfirmationModal.tsx",
        "src\\components\\shared\\NotificationCenter.tsx",
        "src\\components\\shared\\CollectorCard.tsx",
        "src\\components\\shared\\EnterpriseCard.tsx",
        "src\\components\\shared\\ImageGallery.tsx",
        "src\\components\\shared\\RewardCard.tsx",
        "src\\components\\shared\\Toast.tsx",
        "src\\components\\shared\\UserProfileCard.tsx",
        "src\\components\\shared\\UserProfileMenu.tsx",
    ],
    "KIEM-FE-F15": [
        "src\\components\\ui\\Alert.tsx",
        "src\\components\\ui\\Avatar.tsx",
        "src\\components\\ui\\Badge.tsx",
        "src\\components\\ui\\Button.tsx",
        "src\\components\\ui\\Card.tsx",
        "src\\components\\ui\\Dropdown.tsx",
        "src\\components\\ui\\EmptyState.tsx",
        "src\\components\\ui\\Input.tsx",
        "src\\components\\ui\\Modal.tsx",
        "src\\components\\ui\\Pagination.tsx",
        "src\\components\\ui\\Progress.tsx",
        "src\\components\\ui\\Select.tsx",
        "src\\components\\ui\\Spinner.tsx",
        "src\\components\\ui\\Table.tsx",
        "src\\components\\ui\\Toast.tsx",
    ],
}

BACKEND_COVERAGE_GROUPS = {
    "KIEM-4-F01": [
        "WastePlatform.API\\Controllers\\AuthController.cs",
        "WastePlatform.Infrastructure\\Services\\AuthService.cs",
        "WastePlatform.Infrastructure\\Services\\JwtService.cs",
        "WastePlatform.Application\\Auth\\Commands\\RegisterCommand.cs",
        "WastePlatform.Application\\Auth\\Commands\\LoginCommand.cs",
        "WastePlatform.Domain\\Entities\\User.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\UserRepository.cs",
    ],
    "KIEM-5-F02": [
        "WastePlatform.API\\Controllers\\ReportController.cs",
        "WastePlatform.Application\\Reports\\Commands\\CreateReportCommand.cs",
        "WastePlatform.Domain\\Entities\\WasteReport.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\ReportRepository.cs",
        "WastePlatform.Infrastructure\\Services\\LocalFileStorageService.cs",
    ],
    "KIEM-8-F03": [
        "WastePlatform.API\\Controllers\\ReportController.cs",
        "WastePlatform.Application\\Reports\\Commands\\AcceptReportAndCreateTaskCommand.cs",
        "WastePlatform.Application\\Reports\\Commands\\AcceptReportAndCreateTaskCommandHandler.cs",
        "WastePlatform.Application\\Reports\\Commands\\RejectReportCommand.cs",
        "WastePlatform.Application\\Reports\\Commands\\RejectReportCommandHandler.cs",
        "WastePlatform.Domain\\Entities\\WasteReport.cs",
    ],
    "KIEM-10-F04": [
        "WastePlatform.API\\Controllers\\CollectorTaskController.cs",
        "WastePlatform.API\\Controllers\\EnterpriseTaskController.cs",
        "WastePlatform.Domain\\Entities\\CollectionTask.cs",
        "WastePlatform.Domain\\Entities\\TaskStatusLog.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\ReportRepository.cs",
    ],
    "KIEM-16-F05": [
        "WastePlatform.API\\Controllers\\EnterpriseTaskController.cs",
        "WastePlatform.Application\\Tasks\\Commands\\AssignCollectorCommand.cs",
        "WastePlatform.Application\\Tasks\\Commands\\AssignCollectorCommandHandler.cs",
        "WastePlatform.Domain\\Entities\\CollectionTask.cs",
    ],
    "KIEM-13-F06": [
        "WastePlatform.API\\Controllers\\ComplaintsController.cs",
        "WastePlatform.API\\Controllers\\AdminComplaintsController.cs",
        "WastePlatform.Application\\Complaints\\Commands\\CreateComplaintCommand.cs",
        "WastePlatform.Application\\Complaints\\Commands\\EnterpriseRespondToComplaintCommand.cs",
        "WastePlatform.Application\\Complaints\\Commands\\CitizenEscalateComplaintCommand.cs",
        "WastePlatform.Domain\\Entities\\Complaint.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\ComplaintRepository.cs",
    ],
    "KIEM-19-F07": [
        "WastePlatform.API\\Controllers\\NotificationController.cs",
        "WastePlatform.Application\\Services\\NotificationService.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\NotificationRepository.cs",
        "WastePlatform.Infrastructure\\SignalR\\SignalRRealTimeNotifier.cs",
        "WastePlatform.Infrastructure\\SignalR\\TaskHub.cs",
        "WastePlatform.Domain\\Entities\\Notification.cs",
    ],
    "KIEM-20-F08": [
        "WastePlatform.API\\Controllers\\CollectorTaskController.cs",
        "WastePlatform.API\\Middleware\\ValidateUserStatusMiddleware.cs",
        "WastePlatform.Infrastructure\\Services\\JwtService.cs",
        "WastePlatform.Infrastructure\\Services\\LocalFileStorageService.cs",
        "WastePlatform.Domain\\Entities\\CollectionImage.cs",
    ],
    "KIEM-12-F09": [
        "WastePlatform.API\\Controllers\\WasteCategoryController.cs",
        "WastePlatform.Application\\WasteCategories\\Queries\\GetAllCategoriesQuery.cs",
        "WastePlatform.Application\\WasteCategories\\Queries\\GetCategoryByIdQuery.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\WasteCategoryRepository.cs",
        "WastePlatform.Domain\\Entities\\WasteCategory.cs",
    ],
    "KIEM-BVA-F11": [
        "WastePlatform.API\\Controllers\\CollectorTaskController.cs",
        "WastePlatform.Infrastructure\\Services\\LocalFileStorageService.cs",
        "WastePlatform.Domain\\Entities\\CollectionImage.cs",
    ],
    "KIEM-7-F12": [
        "WastePlatform.API\\Controllers\\ComplaintsController.cs",
        "WastePlatform.Application\\Complaints\\Commands\\CreateComplaintCommand.cs",
        "WastePlatform.Domain\\Entities\\Complaint.cs",
    ],
    "KIEM-5-F13": [
        "WastePlatform.API\\Controllers\\ReportController.cs",
        "WastePlatform.Application\\Reports\\Commands\\AcceptReportAndCreateTaskCommand.cs",
        "WastePlatform.Application\\Reports\\Commands\\AcceptReportAndCreateTaskCommandHandler.cs",
        "WastePlatform.Application\\Reports\\Commands\\RejectReportCommand.cs",
        "WastePlatform.Application\\Reports\\Commands\\RejectReportCommandHandler.cs",
        "WastePlatform.Domain\\Entities\\WasteReport.cs",
    ],
}

TECHNIQUE_MAP.update({
    "KIEM-6-F16": "Black-box: Equivalence Partitioning, Error Guessing; Integration: Controller/Repository",
    "KIEM-8-F17": "Black-box: Role-based EP, Authorization Guard, Integration API",
    "KIEM-9-F18": "Black-box: Equivalence Partitioning, Empty Data/Error Guessing",
    "KIEM-10-F19": "Black-box: Public API Smoke + EP for analytics filters",
    "KIEM-13-F20": "Black-box: Citizen profile EP, Validation/Error Guessing",
    "KIEM-14-F21": "Black-box + E2E: Collector role workflow, Authorization Guard",
    "KIEM-15-F22": "Black-box: CollectorTask workflow, BVA for evidence upload",
    "KIEM-17-F23": "Black-box: Enterprise collector CRUD, Reward Rules EP/BVA",
    "KIEM-18-F24": "White-box/Domain: CollectionTask State Transition + Branch Coverage",
    "KIEM-21-F25": "White-box + Integration: JWT, RBAC, Middleware Branch Coverage",
    "KIEM-22-F26": "Black-box + White-box: AuditLog, Exception/Error Path Coverage",
    "KIEM-23-F27": "Black-box: Search, Pagination, Filter Equivalence Classes",
    "KIEM-E2E-F28": "E2E: Smoke/Auth Navigation (CodeceptJS + Playwright)",
    "KIEM-E2E-F29": "E2E: Role Dashboard Guards (Admin/Citizen/Settings)",
    "KIEM-E2E-F30": "E2E: Enterprise + Collector Operational Workflow",
})

ROLE_MAP.update({
    "KIEM-6-F16": "Citizen / Enterprise / Admin",
    "KIEM-8-F17": "Admin",
    "KIEM-9-F18": "Admin / Enterprise",
    "KIEM-10-F19": "Public / Guest",
    "KIEM-13-F20": "Citizen",
    "KIEM-14-F21": "Collector",
    "KIEM-15-F22": "Collector / Enterprise",
    "KIEM-17-F23": "Enterprise",
    "KIEM-18-F24": "Enterprise / Collector",
    "KIEM-21-F25": "Admin / Enterprise / Citizen / Collector",
    "KIEM-22-F26": "Admin / API Client",
    "KIEM-23-F27": "Admin / Enterprise / Citizen",
    "KIEM-E2E-F28": "Guest / Citizen",
    "KIEM-E2E-F29": "Admin / Citizen",
    "KIEM-E2E-F30": "Enterprise / Collector",
})

BACKEND_COVERAGE_GROUPS.update({
    "KIEM-6-F16": [
        "WastePlatform.API\\Controllers\\NotificationController.cs",
        "WastePlatform.Application\\Services\\NotificationService.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\NotificationRepository.cs",
        "WastePlatform.Domain\\Entities\\Notification.cs",
        "WastePlatform.Infrastructure\\SignalR\\SignalRRealTimeNotifier.cs",
    ],
    "KIEM-8-F17": [
        "WastePlatform.API\\Controllers\\AdminUsersController.cs",
        "WastePlatform.API\\Controllers\\AdminEnterpriseController.cs",
        "WastePlatform.API\\Controllers\\AdminComplaintsController.cs",
        "WastePlatform.Application\\Admin\\Users\\Commands\\CreateUserCommand.cs",
        "WastePlatform.Application\\Admin\\Users\\Commands\\ToggleUserStatusCommand.cs",
        "WastePlatform.Application\\Admin\\Users\\Commands\\UpdateUserRoleCommand.cs",
    ],
    "KIEM-9-F18": [
        "WastePlatform.API\\Controllers\\AdminAnalyticsController.cs",
        "WastePlatform.API\\Controllers\\EnterpriseAnalyticsController.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\AnalyticsRepository.cs",
        "WastePlatform.Application\\Admin\\Analytics\\Queries\\GetAnalyticsOverviewQuery.cs",
        "WastePlatform.Application\\Admin\\Analytics\\Queries\\GetAnalyticsSummaryQuery.cs",
        "WastePlatform.Application\\Admin\\Analytics\\Queries\\GetReportAnalyticsQuery.cs",
        "WastePlatform.Application\\Admin\\Analytics\\Queries\\GetUserAnalyticsQuery.cs",
        "WastePlatform.Application\\Admin\\Analytics\\Queries\\GetWasteAnalyticsQuery.cs",
    ],
    "KIEM-10-F19": [
        "WastePlatform.API\\Controllers\\PublicAnalyticsController.cs",
        "WastePlatform.API\\Controllers\\HealthController.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\AnalyticsRepository.cs",
    ],
    "KIEM-13-F20": [
        "WastePlatform.API\\Controllers\\CitizenController.cs",
        "WastePlatform.Application\\Citizens\\Profile\\Queries\\GetProfileQueryHandler.cs",
        "WastePlatform.Application\\Citizens\\Profile\\Commands\\UpdateProfileCommandHandler.cs",
        "WastePlatform.Application\\Citizens\\Profile\\DTOs\\ProfileDto.cs",
        "WastePlatform.Application\\Citizens\\Profile\\DTOs\\UpdateProfileDto.cs",
    ],
    "KIEM-14-F21": [
        "WastePlatform.API\\Controllers\\CollectorController.cs",
        "WastePlatform.API\\Controllers\\CollectorTaskController.cs",
        "WastePlatform.Domain\\Entities\\Collector.cs",
        "WastePlatform.Domain\\Entities\\CollectionTask.cs",
    ],
    "KIEM-15-F22": [
        "WastePlatform.API\\Controllers\\CollectorTaskController.cs",
        "WastePlatform.API\\Controllers\\EnterpriseTaskController.cs",
        "WastePlatform.Application\\Tasks\\Commands\\AssignCollectorCommand.cs",
        "WastePlatform.Application\\Tasks\\Commands\\AssignCollectorCommandHandler.cs",
        "WastePlatform.Domain\\Entities\\CollectionTask.cs",
        "WastePlatform.Domain\\Entities\\CollectionImage.cs",
    ],
    "KIEM-17-F23": [
        "WastePlatform.API\\Controllers\\EnterpriseCollectorController.cs",
        "WastePlatform.API\\Controllers\\EnterpriseRewardRuleController.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\RewardPointsRepository.cs",
        "WastePlatform.Application\\Rewards\\Commands\\CreateRewardPointsCommand.cs",
        "WastePlatform.Application\\Rewards\\Commands\\CreateRewardPointsCommandHandler.cs",
        "WastePlatform.Domain\\Entities\\RewardRule.cs",
        "WastePlatform.Domain\\Entities\\RewardPoints.cs",
    ],
    "KIEM-18-F24": [
        "WastePlatform.Domain\\Entities\\CollectionTask.cs",
        "WastePlatform.Domain\\Entities\\TaskStatusLog.cs",
        "WastePlatform.Domain\\Events\\TaskStatusChangedEvent.cs",
    ],
    "KIEM-21-F25": [
        "WastePlatform.API\\Middleware\\ValidateUserStatusMiddleware.cs",
        "WastePlatform.Infrastructure\\Services\\JwtService.cs",
        "WastePlatform.API\\Controllers\\AuthController.cs",
        "WastePlatform.Domain\\Entities\\User.cs",
    ],
    "KIEM-22-F26": [
        "WastePlatform.API\\Controllers\\AuditLogAndErrorPathTests.cs",
        "WastePlatform.Domain\\Entities\\AuditLog.cs",
        "WastePlatform.API\\Controllers\\HealthController.cs",
        "WastePlatform.API\\Middleware\\ValidateUserStatusMiddleware.cs",
    ],
    "KIEM-23-F27": [
        "WastePlatform.Application\\Reports\\Queries\\GetAllReportsQuery.cs",
        "WastePlatform.Application\\Reports\\Queries\\GetEnterpriseReportsQuery.cs",
        "WastePlatform.Application\\Complaints\\Queries\\GetCitizenComplaintsQuery.cs",
        "WastePlatform.Application\\Complaints\\Queries\\GetEnterpriseComplaintsQuery.cs",
        "WastePlatform.Application\\Admin\\Complaints\\Queries\\GetComplaintsQuery.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\ReportRepository.cs",
        "WastePlatform.Infrastructure\\Persistence\\Repositories\\ComplaintRepository.cs",
    ],
})


def safe_percent(covered, total):
    if total == 0:
        return 100.0
    return round((covered / total) * 100, 2)


def format_coverage(metric):
    if not metric:
        return "Not instrumented"
    covered, total = metric
    return f"{safe_percent(covered, total)}% ({covered}/{total})"


def load_frontend_coverage_summary():
    if not os.path.exists(FRONTEND_COVERAGE_SUMMARY):
        return {}
    with open(FRONTEND_COVERAGE_SUMMARY, "r", encoding="utf-8") as f:
        return json.load(f)


def load_backend_coverage_json():
    if not os.path.exists(BACKEND_COVERAGE_JSON):
        return {}
    with open(BACKEND_COVERAGE_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def aggregate_frontend_coverage(summary, suffixes):
    totals = {
        "line": [0, 0],
        "branch": [0, 0],
        "function": [0, 0],
        "statement": [0, 0],
    }
    normalized_suffixes = [s.replace("/", "\\").lower() for s in suffixes]
    for path, metrics in summary.items():
        normalized_path = path.replace("/", "\\").lower()
        if not any(normalized_path.endswith(sfx) for sfx in normalized_suffixes):
            continue
        totals["line"][0] += metrics.get("lines", {}).get("covered", 0)
        totals["line"][1] += metrics.get("lines", {}).get("total", 0)
        totals["branch"][0] += metrics.get("branches", {}).get("covered", 0)
        totals["branch"][1] += metrics.get("branches", {}).get("total", 0)
        totals["function"][0] += metrics.get("functions", {}).get("covered", 0)
        totals["function"][1] += metrics.get("functions", {}).get("total", 0)
        totals["statement"][0] += metrics.get("statements", {}).get("covered", 0)
        totals["statement"][1] += metrics.get("statements", {}).get("total", 0)
    return {k: tuple(v) for k, v in totals.items()}


def aggregate_backend_coverage(summary, suffixes):
    totals = {
        "line": [0, 0],
        "branch": [0, 0],
        "function": [0, 0],
        "statement": [0, 0],
    }
    normalized_suffixes = [s.replace("/", "\\").lower() for s in suffixes]
    matched_files = 0

    for module_files in summary.values():
        for path, classes in module_files.items():
            normalized_path = path.replace("/", "\\").lower()
            if not any(normalized_path.endswith(sfx) for sfx in normalized_suffixes):
                continue
            matched_files += 1

            file_lines = {}
            file_branches = []
            methods_total = 0
            methods_covered = 0

            for methods in classes.values():
                for method_data in methods.values():
                    methods_total += 1
                    method_lines = method_data.get("Lines", {})
                    if any(count > 0 for count in method_lines.values()):
                        methods_covered += 1

                    for line, count in method_lines.items():
                        file_lines[line] = max(file_lines.get(line, 0), count)

                    file_branches.extend(method_data.get("Branches", []))

            totals["line"][0] += sum(1 for count in file_lines.values() if count > 0)
            totals["line"][1] += len(file_lines)
            totals["statement"][0] += sum(1 for count in file_lines.values() if count > 0)
            totals["statement"][1] += len(file_lines)
            totals["function"][0] += methods_covered
            totals["function"][1] += methods_total
            totals["branch"][0] += sum(1 for entry in file_branches if entry.get("Hits", 0) > 0)
            totals["branch"][1] += len(file_branches)

    if matched_files == 0:
        return {}
    return {k: tuple(v) for k, v in totals.items()}


def get_test_design_coverage(func):
    n_utcid = len(func["utcids"])
    lack = max(0, 7 - n_utcid)
    total_required = n_utcid + lack
    return (n_utcid, total_required)


def build_coverage_info(func, frontend_summary=None, backend_summary=None):
    code = func["code"]
    design_metric = get_test_design_coverage(func)
    info = {
        "test_design": design_metric,
        "blackbox": design_metric,
        "line": None,
        "branch": None,
        "statement": None,
        "condition": None,
        "path": None,
        "source": "Test design matrix trong file generate_unittest_excel.py",
        "note": "Black-box coverage tính theo số UTCID đã thiết kế so với baseline tối thiểu 7 test case/function.",
    }

    if code in BACKEND_COVERAGE_GROUPS:
        summary = backend_summary if backend_summary is not None else load_backend_coverage_json()
        aggregate = aggregate_backend_coverage(summary, BACKEND_COVERAGE_GROUPS[code])
        if aggregate:
            info.update(aggregate)
            info["source"] = "backend/coverage/backend/coverage.json (dotnet test + coverlet)"
            info["note"] = "Line/branch coverage được cộng từ các source files backend liên quan trực tiếp tới function."

    if code in WHITEBOX_COVERAGE_MAP:
        documented_whitebox = WHITEBOX_COVERAGE_MAP[code]
        info["condition"] = documented_whitebox["condition"]
        info["path"] = documented_whitebox["path"]
        info["note"] = info["note"] + " " + documented_whitebox["note"]
        info["blackbox"] = design_metric
        return info

    if code in FRONTEND_COVERAGE_GROUPS:
        summary = frontend_summary if frontend_summary is not None else load_frontend_coverage_summary()
        aggregate = aggregate_frontend_coverage(summary, FRONTEND_COVERAGE_GROUPS[code])
        info.update(aggregate)
        info["source"] = "frontend/coverage/coverage-summary.json (Vitest --coverage)"
        info["note"] = "Line/branch coverage được tính từ các component source tương ứng với function sheet."

    return info


def build_test_case_detail(func, ui):
    conditions_text = []
    for cg in func["conditions"]:
        for item in cg["items"]:
            if ui in item["marks"]:
                conditions_text.append(f"[{cg['group']}] {item['label']}")

    ret_codes = [r["code"] for r in func["returns"] if ui in r["marks"]]
    log_msgs = [
        lg.get("msg", lg.get("label", ""))
        for lg in func["logs"]
        if ui in lg["marks"]
    ]
    result = func["results"][ui]
    test_type = {
        "N": "Normal",
        "A": "Abnormal",
        "B": "Boundary",
    }.get(result["type"], result["type"])

    return (
        f"Mục tiêu: {func['test_req']}\n"
        f"Input/Precondition:\n" + ("\n".join(f"- {c}" for c in conditions_text) or "- N/A") + "\n"
        f"Kỹ thuật: {TECHNIQUE_MAP.get(func['code'], '')}\n"
        f"Loại test: {test_type}\n"
        f"Expected: {', '.join(ret_codes) if ret_codes else 'Theo assertion trong test'}\n"
        f"Actual/Log: {'; '.join(log_msgs) if log_msgs else 'Không có log lỗi'}\n"
        f"Result: {'Passed' if result['pf'] == 'P' else 'Failed'}"
    )

# ─── Định nghĩa dữ liệu các Function ────────────────────────────────
#
# Mỗi function dict có:
#   code        : mã function
#   name        : tên function
#   created_by  : tên người tạo
#   executed_by : tên người chạy
#   lines_of_code: dòng code ước tính
#   jira_ticket : KIEM-xx
#   test_req    : yêu cầu kiểm thử
#   utcids      : danh sách ID cột test (UTC01, UTC02...)
#   conditions  : [{"group": "Tên nhóm", "items": [{"label": "giá trị", "marks": [0,1,...]}]}]
#                 marks: list index UTCID (0-based) nào có "O"
#   returns     : [{"code": "200", "marks": [...]}]
#   exceptions  : [{"msg": "...", "marks": [...]}] hoặc []
#   logs        : [{"msg": "...", "marks": [...]}]
#   results     : [{"type":"N"|"A"|"B", "pf":"P"|"F", "date":TODAY, "defect":""}]
#                 (1 entry per UTCID)

FUNCTIONS = [

    # ─── FUNCTION 1: Auth — Đăng ký & Đăng nhập ─────────────────────
    {
        "code": "KIEM-4-F01",
        "name": "Xác thực người dùng (Đăng ký / Đăng nhập)",
        "created_by": "Nguyễn Chí Trung",
        "executed_by": "Nguyễn Chí Trung",
        "lines_of_code": 350,
        "jira_ticket": "KIEM-4",
        "test_req": (
            "Hệ thống cho phép đăng ký tài khoản mới (Citizen/Enterprise); "
            "đăng nhập trả về JWT token hợp lệ; "
            "đăng nhập sai mật khẩu trả về 401."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05"],
        "conditions": [
            {
                "group": "email",
                "items": [
                    {"label": "citizen@gmail.com (hợp lệ)", "marks": [0, 1]},
                    {"label": "null / rỗng",                "marks": [2]},
                    {"label": "định dạng sai (abc@)",        "marks": [3]},
                    {"label": "email trùng (duplicate)",     "marks": [4]},
                ]
            },
            {
                "group": "password",
                "items": [
                    {"label": "*Pass@123 (hợp lệ)",         "marks": [0]},
                    {"label": "*sai mật khẩu",              "marks": [3]},
                    {"label": "null / rỗng",                "marks": [2]},
                ]
            },
            {
                "group": "role",
                "items": [
                    {"label": "Citizen",                    "marks": [0, 1, 2, 3]},
                    {"label": "Enterprise",                 "marks": [4]},
                ]
            },
        ],
        "returns": [
            {"code": "200 / 201",  "marks": [0, 1]},
            {"code": "400",        "marks": [2, 3]},
            {"code": "401",        "marks": [3]},
            {"code": "409",        "marks": [4]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Tài khoản tạo thành công",      "marks": [0, 1]},
            {"msg": "Email hoặc mật khẩu không đúng","marks": [3]},
            {"msg": "Email không được để trống",     "marks": [2]},
            {"msg": "Email đã tồn tại trong hệ thống","marks": [4]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 2: Create Waste Report ────────────────────────────
    {
        "code": "KIEM-5-F02",
        "name": "Tạo báo cáo rác thải (Citizen)",
        "created_by": "Minh Phụng",
        "executed_by": "Minh Phụng",
        "lines_of_code": 220,
        "jira_ticket": "KIEM-5",
        "test_req": (
            "Citizen tạo báo cáo với hình ảnh, toạ độ hợp lệ; "
            "hệ thống từ chối khi thiếu ảnh hoặc toạ độ ngoài phạm vi."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05", "UTC06"],
        "conditions": [
            {
                "group": "images",
                "items": [
                    {"label": "≥1 ảnh hợp lệ",             "marks": [0, 1, 5]},
                    {"label": "danh sách rỗng ([])",         "marks": [2]},
                    {"label": "null",                        "marks": [3]},
                ]
            },
            {
                "group": "latitude",
                "items": [
                    {"label": "10.776889 (hợp lệ)",          "marks": [0, 1]},
                    {"label": "-90 (biên âm)",               "marks": [4]},
                    {"label": "90 (biên dương)",             "marks": [5]},
                    {"label": "-91 / 91 (ngoài phạm vi)",    "marks": [3]},
                ]
            },
            {
                "group": "longitude",
                "items": [
                    {"label": "106.700981 (hợp lệ)",         "marks": [0, 1]},
                    {"label": "-180 / 180 (biên)",           "marks": [4, 5]},
                    {"label": "ngoài -180..180",             "marks": [3]},
                ]
            },
            {
                "group": "wasteCategoryId",
                "items": [
                    {"label": "ID hợp lệ (tồn tại)",        "marks": [0, 4, 5]},
                    {"label": "ID không tồn tại",           "marks": [2]},
                ]
            },
        ],
        "returns": [
            {"code": "200 / 201", "marks": [0, 1, 4, 5]},
            {"code": "400",       "marks": [2, 3]},
        ],
        "exceptions": [
            {"msg": "ArgumentException: Images required", "marks": [2, 3]},
            {"msg": "ArgumentException: Invalid coordinates", "marks": [3]},
        ],
        "logs": [
            {"msg": "Báo cáo tạo thành công, ID: {guid}",  "marks": [0, 1, 4, 5]},
            {"msg": "Thiếu ảnh hoặc danh sách rỗng",       "marks": [2, 3]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 3: Accept/Reject Report ───────────────────────────
    {
        "code": "KIEM-8-F03",
        "name": "Duyệt / Từ chối báo cáo (Enterprise/Admin)",
        "created_by": "Đăng",
        "executed_by": "Đăng",
        "lines_of_code": 180,
        "jira_ticket": "KIEM-8",
        "test_req": (
            "Enterprise/Admin có thể accept hoặc reject báo cáo ở trạng thái Pending; "
            "hệ thống từ chối thao tác khi báo cáo đã được xử lý."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05"],
        "conditions": [
            {
                "group": "Trạng thái báo cáo (report status)",
                "items": [
                    {"label": "Pending",          "marks": [0, 1]},
                    {"label": "Accepted",         "marks": [2]},
                    {"label": "Rejected",         "marks": [3]},
                    {"label": "Collected",        "marks": [4]},
                    {"label": "Không tồn tại",   "marks": [1]},
                ]
            },
            {
                "group": "Hành động",
                "items": [
                    {"label": "Accept (chấp nhận)", "marks": [0, 2, 3, 4]},
                    {"label": "Reject (từ chối)",   "marks": [1]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0, 1]},
            {"code": "400",  "marks": [2, 3, 4]},
        ],
        "exceptions": [
            {"msg": "InvalidOperationException: Report is not Pending", "marks": [2, 3, 4]},
        ],
        "logs": [
            {"msg": "Báo cáo được chấp nhận thành công",  "marks": [0]},
            {"msg": "Báo cáo bị từ chối",                 "marks": [1]},
            {"msg": "Không thể thay đổi trạng thái hiện tại", "marks": [2, 3, 4]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 4: Collection Task — State Transition ─────────────
    {
        "code": "KIEM-10-F04",
        "name": "Vòng đời nhiệm vụ thu gom (CollectionTask State)",
        "created_by": "Thanh Duy",
        "executed_by": "Thanh Duy",
        "lines_of_code": 250,
        "jira_ticket": "KIEM-10",
        "test_req": (
            "CollectionTask chuyển trạng thái đúng quy tắc domain: "
            "Assigned → OnTheWay → Collected; "
            "không cho phép chuyển trạng thái sai thứ tự."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05", "UTC06"],
        "conditions": [
            {
                "group": "Trạng thái hiện tại",
                "items": [
                    {"label": "Assigned",          "marks": [0, 1]},
                    {"label": "OnTheWay",           "marks": [2, 3]},
                    {"label": "Collected",          "marks": [4]},
                    {"label": "Không tồn tại (ID sai)", "marks": [5]},
                ]
            },
            {
                "group": "Hành động",
                "items": [
                    {"label": "SetOnTheWay()",      "marks": [0, 2, 4]},
                    {"label": "Complete()",          "marks": [1, 3]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0, 3]},
            {"code": "400",  "marks": [1, 2, 4, 5]},
            {"code": "404",  "marks": [5]},
        ],
        "exceptions": [
            {"msg": "InvalidOperationException: SetOnTheWay khi không phải Assigned", "marks": [2, 4]},
            {"msg": "InvalidOperationException: Complete khi không phải OnTheWay",    "marks": [1]},
        ],
        "logs": [
            {"msg": "Task chuyển sang OnTheWay",    "marks": [0]},
            {"msg": "Task hoàn thành (Collected)",  "marks": [3]},
            {"msg": "Không thể thực hiện thao tác", "marks": [1, 2, 4]},
            {"msg": "Task không tồn tại",           "marks": [5]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 5: Assign Collector ───────────────────────────────
    {
        "code": "KIEM-16-F05",
        "name": "Enterprise giao nhiệm vụ cho Collector",
        "created_by": "Nguyễn Chí Trung",
        "executed_by": "Nguyễn Chí Trung",
        "lines_of_code": 200,
        "jira_ticket": "KIEM-16",
        "test_req": (
            "Enterprise gán collector hợp lệ vào task; "
            "hệ thống từ chối khi collector không tồn tại hoặc không thuộc enterprise."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04"],
        "conditions": [
            {
                "group": "Enterprise",
                "items": [
                    {"label": "Enterprise hợp lệ (đã xác thực)", "marks": [0, 1, 2]},
                    {"label": "Enterprise không hợp lệ",          "marks": [3]},
                ]
            },
            {
                "group": "CollectorId",
                "items": [
                    {"label": "Collector thuộc enterprise",       "marks": [0]},
                    {"label": "Collector không thuộc enterprise", "marks": [1]},
                    {"label": "CollectorId không tồn tại",        "marks": [2]},
                ]
            },
            {
                "group": "Task Status",
                "items": [
                    {"label": "Assigned (chưa có collector)",     "marks": [0, 1]},
                    {"label": "Task đã có collector",             "marks": [3]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0]},
            {"code": "400",  "marks": [1, 2, 3]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Collector gán thành công, thông báo realtime gửi đến Citizen", "marks": [0]},
            {"msg": "Collector không hợp lệ hoặc không thuộc enterprise",           "marks": [1, 2]},
            {"msg": "Không có quyền thực hiện",                                     "marks": [3]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 6: Complaints ──────────────────────────────────────
    {
        "code": "KIEM-13-F06",
        "name": "Quản lý khiếu nại (Complaints)",
        "created_by": "Đăng",
        "executed_by": "Đăng",
        "lines_of_code": 300,
        "jira_ticket": "KIEM-13",
        "test_req": (
            "Citizen tạo khiếu nại liên kết đến report; "
            "Admin resolve hoặc reject khiếu nại; "
            "hệ thống từ chối khi nội dung rỗng."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05"],
        "conditions": [
            {
                "group": "content",
                "items": [
                    {"label": "Nội dung hợp lệ (≤2000 ký tự)",   "marks": [0, 1]},
                    {"label": "Rỗng / null",                       "marks": [2]},
                    {"label": "Vượt 2000 ký tự",                   "marks": [3]},
                ]
            },
            {
                "group": "reportId",
                "items": [
                    {"label": "Report tồn tại (status Accepted/Collected)", "marks": [0, 1]},
                    {"label": "Report không tồn tại",                       "marks": [4]},
                    {"label": "Report trạng thái Pending",                  "marks": [3]},
                ]
            },
            {
                "group": "Hành động Admin",
                "items": [
                    {"label": "Resolve (giải quyết)", "marks": [1]},
                    {"label": "Reject (từ chối)",     "marks": [2]},
                ]
            },
        ],
        "returns": [
            {"code": "200 / 201", "marks": [0, 1]},
            {"code": "400",       "marks": [2, 3, 4]},
        ],
        "exceptions": [
            {"msg": "ArgumentException: Content không được rỗng",     "marks": [2]},
            {"msg": "InvalidOperationException: Report chưa được xử lý", "marks": [3]},
        ],
        "logs": [
            {"msg": "Khiếu nại tạo thành công",      "marks": [0]},
            {"msg": "Khiếu nại đã được giải quyết",  "marks": [1]},
            {"msg": "Nội dung không hợp lệ",         "marks": [2, 3]},
            {"msg": "Report không tồn tại",          "marks": [4]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 7: Notifications ──────────────────────────────────
    {
        "code": "KIEM-19-F07",
        "name": "Hệ thống thông báo (Notifications + SignalR)",
        "created_by": "Nguyễn Chí Trung",
        "executed_by": "Nguyễn Chí Trung",
        "lines_of_code": 280,
        "jira_ticket": "KIEM-19",
        "test_req": (
            "Thông báo được lưu vào DB và gửi realtime qua SignalR khi tạo báo cáo / reject / escalate; "
            "Citizen xem danh sách thông báo và đánh dấu đã đọc."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04"],
        "conditions": [
            {
                "group": "Sự kiện kích hoạt",
                "items": [
                    {"label": "Báo cáo được tạo",          "marks": [0]},
                    {"label": "Báo cáo bị reject",          "marks": [1]},
                    {"label": "Khiếu nại escalate → Admin", "marks": [2]},
                    {"label": "Citizen xem / đánh dấu đọc","marks": [3]},
                ]
            },
            {
                "group": "citizenId",
                "items": [
                    {"label": "Citizen hợp lệ",       "marks": [0, 1, 2, 3]},
                    {"label": "CitizenId null / thiếu","marks": [3]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0, 1, 2, 3]},
            {"code": "401",  "marks": [3]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Notification saved + SignalR push",        "marks": [0, 1]},
            {"msg": "Admin notification (no realtime push)",    "marks": [2]},
            {"msg": "Danh sách thông báo trả về (paged)",      "marks": [3]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 8: File Upload & Security ─────────────────────────
    {
        "code": "KIEM-20-F08",
        "name": "Upload ảnh bằng chứng & Bảo mật JWT",
        "created_by": "Minh Phụng",
        "executed_by": "Minh Phụng",
        "lines_of_code": 310,
        "jira_ticket": "KIEM-20",
        "test_req": (
            "Collector upload ảnh bằng chứng (jpg/png/webp ≤10MB); "
            "JWT token hợp lệ cho phép truy cập endpoint được bảo vệ; "
            "Token sai/hết hạn trả về 401."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03", "UTC04", "UTC05"],
        "conditions": [
            {
                "group": "File upload",
                "items": [
                    {"label": "*.jpg hợp lệ (≤10MB)",     "marks": [0]},
                    {"label": "*.png / *.webp",           "marks": [1]},
                    {"label": "File rỗng (0 byte)",       "marks": [2]},
                    {"label": "Đuôi file không cho phép (.exe)", "marks": [3]},
                    {"label": "Vượt 10MB",                "marks": [4]},
                ]
            },
            {
                "group": "JWT Token",
                "items": [
                    {"label": "Token hợp lệ",             "marks": [0, 1]},
                    {"label": "Không có token",           "marks": [2]},
                    {"label": "Token sai chữ ký",         "marks": [3]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0, 1]},
            {"code": "400",  "marks": [2, 3, 4]},
            {"code": "401",  "marks": [2, 3]},
        ],
        "exceptions": [
            {"msg": "ArgumentException: File empty",               "marks": [2]},
            {"msg": "InvalidOperationException: Extension not allowed", "marks": [3]},
            {"msg": "InvalidOperationException: File too large",   "marks": [4]},
        ],
        "logs": [
            {"msg": "File lưu thành công, trả về filename GUID",  "marks": [0, 1]},
            {"msg": "Lỗi upload: file rỗng hoặc không hợp lệ",   "marks": [2, 3, 4]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 9: Waste Category ─────────────────────────────────
    {
        "code": "KIEM-12-F09",
        "name": "Danh mục loại rác thải (Waste Category)",
        "created_by": "Nguyễn Hoàng Phụng",
        "executed_by": "Nguyễn Hoàng Phụng",
        "lines_of_code": 120,
        "jira_ticket": "KIEM-12",
        "test_req": (
            "API trả về danh sách loại rác thải có sắp xếp theo tên; "
            "lấy theo ID trả về đúng hoặc 404 nếu không tồn tại."
        ),
        "utcids": ["UTC01", "UTC02", "UTC03"],
        "conditions": [
            {
                "group": "Request",
                "items": [
                    {"label": "GET /api/waste-categories (lấy tất cả)", "marks": [0]},
                    {"label": "GET /api/waste-categories/{id} (hợp lệ)","marks": [1]},
                    {"label": "GET /api/waste-categories/{id} (không tồn tại)", "marks": [2]},
                ]
            },
            {
                "group": "DB State",
                "items": [
                    {"label": "Có ≥1 category trong DB",  "marks": [0, 1]},
                    {"label": "DB rỗng",                  "marks": [0]},
                ]
            },
        ],
        "returns": [
            {"code": "200",  "marks": [0, 1]},
            {"code": "404",  "marks": [2]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Danh sách category sắp xếp theo tên", "marks": [0]},
            {"msg": "Category tìm thấy theo ID",           "marks": [1]},
            {"msg": "Category không tồn tại",              "marks": [2]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 10: E2E — Citizen Report Flow ──────────────────────
    {
        "code": "KIEM-FE-F10",
        "name": "E2E: Citizen đăng ký và tạo báo cáo (CodeceptJS)",
        "created_by": "Nguyễn Chí Trung",
        "executed_by": "Nguyễn Chí Trung",
        "lines_of_code": 140,
        "jira_ticket": "KIEM-FE",
        "test_req": (
            "Citizen đăng ký thành công, điều hướng đến trang tạo báo cáo; "
            "form validate bắt buộc khi thiếu trường; "
            "truy cập /create-report khi chưa login bị chặn."
        ),
        "utcids": ["E2E-01", "E2E-02", "E2E-03"],
        "conditions": [
            {
                "group": "Precondition",
                "items": [
                    {"label": "Backend chạy tại localhost:8080",      "marks": [0, 1, 2]},
                    {"label": "Frontend chạy tại localhost:3000",     "marks": [0, 1, 2]},
                    {"label": "Citizen account seeded (citizen@test.waste)", "marks": [1, 2]},
                ]
            },
            {
                "group": "Hành động",
                "items": [
                    {"label": "Đăng ký tài khoản Citizen mới",       "marks": [0]},
                    {"label": "Đăng nhập → vào trang Create Report", "marks": [1]},
                    {"label": "Truy cập /create-report chưa login",  "marks": [2]},
                ]
            },
        ],
        "returns": [
            {"code": "Redirect /create-report",   "marks": [0, 1]},
            {"code": "Redirect /login (guard)",   "marks": [2]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Trang create-report hiển thị sau login (Citizen đã đăng nhập)", "marks": [0, 1]},
            {"msg": "Guard chặn, redirect về /login (chưa xác thực)",              "marks": [2]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},
        ],
    },

    # ─── FUNCTION 11: BVA — Images Upload Constraints (KIEM-26/29) ──
    {
        "code": "KIEM-BVA-F11",
        "name": "Kiểm thử BVA — Ràng buộc Upload Ảnh (1≤images≤5)",
        "created_by": "Nguyễn Hoàng Phụng",
        "executed_by": "Nguyễn Hoàng Phụng",
        "lines_of_code": 180,
        "jira_ticket": "KIEM-26",
        "test_req": (
            "Áp dụng BVA (Standard + Robustness) theo giáo trình Ch.4: "
            "Upload ảnh bằng chứng phải thỏa min=1, max=5 ảnh. "
            "KIEM-26: thiếu validation ảnh bắt buộc. "
            "KIEM-29: thiếu validation tối đa 5 ảnh."
        ),
        "utcids": [
            "BVA-01",   # images=0 (dưới min, invalid)
            "BVA-02",   # images=1 (đúng min, valid)
            "BVA-03",   # images=2 (min+1, valid)
            "BVA-04",   # images=3 (nominal, valid)
            "BVA-05",   # images=4 (max-1, valid)
            "BVA-06",   # images=5 (đúng max, valid)
            "BVA-07",   # images=6 (vượt max, invalid — KIEM-29 bug)
            "BVA-08",   # images=null (không gửi field — KIEM-26 bug)
        ],
        "conditions": [
            {
                "group": "Số lượng ảnh (images count)",
                "items": [
                    {"label": "0 ảnh (dưới min — invalid)",       "marks": [0]},
                    {"label": "1 ảnh (đúng min — valid BVA)",      "marks": [1]},
                    {"label": "2 ảnh (min+1 — valid BVA)",         "marks": [2]},
                    {"label": "3 ảnh (giữa — nominal)",            "marks": [3]},
                    {"label": "4 ảnh (max-1 — valid BVA)",         "marks": [4]},
                    {"label": "5 ảnh (đúng max — valid BVA)",      "marks": [5]},
                    {"label": "6 ảnh (vượt max — invalid BVA)",    "marks": [6]},
                    {"label": "null / không có field",             "marks": [7]},
                ]
            },
            {
                "group": "File type",
                "items": [
                    {"label": "*.jpg / *.png / *.webp (hợp lệ)", "marks": [1, 2, 3, 4, 5]},
                    {"label": "Loại file không hợp lệ (.exe)",   "marks": [0]},
                ]
            },
            {
                "group": "JWT Auth",
                "items": [
                    {"label": "Token hợp lệ (Citizen đã đăng nhập)", "marks": [0, 1, 2, 3, 4, 5, 6, 7]},
                ]
            },
        ],
        "returns": [
            {"code": "201 Created",       "marks": [1, 2, 3, 4, 5]},
            {"code": "400 Bad Request",   "marks": [0, 6, 7]},
        ],
        "exceptions": [
            {"msg": "ArgumentException: Cần ít nhất 1 ảnh (KIEM-26 bug fix)", "marks": [0, 7]},
            {"msg": "ArgumentException: Tối đa 5 ảnh (KIEM-29 bug fix)",     "marks": [6]},
        ],
        "logs": [
            {"msg": "Ảnh upload thành công (1-5 ảnh)",              "marks": [1, 2, 3, 4, 5]},
            {"msg": "Lỗi: Thiếu ảnh bắt buộc (Bug KIEM-26)",        "marks": [0, 7]},
            {"msg": "Lỗi: Vượt quá số ảnh tối đa 5 (Bug KIEM-29)", "marks": [6]},
        ],
        "results": [
            {"type": "A", "pf": "P", "date": TODAY, "defect": "KIEM-26"},  # BVA-01: 0 images — bug fixed, validation throws ArgEx
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},         # BVA-02: 1 image
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},         # BVA-03: 2 images
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},         # BVA-04: 3 images
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},         # BVA-05: 4 images
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},         # BVA-06: 5 images
            {"type": "A", "pf": "P", "date": TODAY, "defect": "KIEM-29"},  # BVA-07: 6 images — bug fixed, validation throws ArgEx
            {"type": "A", "pf": "P", "date": TODAY, "defect": "KIEM-26"},  # BVA-08: null — bug fixed, validation throws ArgEx
        ],
    },

    # ─── FUNCTION 12: Decision Table — Complaint Creation (KIEM-7) ──
    {
        "code": "KIEM-7-F12",
        "name": "Decision Table — Tạo khiếu nại (Complaints)",
        "created_by": "Thanh Duy",
        "executed_by": "Thanh Duy",
        "lines_of_code": 300,
        "jira_ticket": "KIEM-7",
        "test_req": (
            "Áp dụng Decision Table Testing (Ch.4 §IV.3): "
            "Xác định tất cả tổ hợp điều kiện đầu vào khi tạo khiếu nại. "
            "Conditions: Content hợp lệ/rỗng × Report tồn tại/không × Report status Valid/Pending."
        ),
        "utcids": [
            "DT-01",  # Content Valid + Report Valid Accepted → 201
            "DT-02",  # Content Valid + Report Valid Pending → 400 (InvalidOp)
            "DT-03",  # Content Valid + Report không tồn tại → 400 (ArgEx)
            "DT-04",  # Content Valid + No reportId (direct) → 201
            "DT-05",  # Content rỗng/null + bất kỳ → 400 (ArgEx)
            "DT-06",  # Content > 2000 chars → 400 (ArgEx)
        ],
        "conditions": [
            {
                "group": "Content (nội dung khiếu nại)",
                "items": [
                    {"label": "Hợp lệ (1-2000 ký tự)",          "marks": [0, 1, 2, 3]},
                    {"label": "Rỗng / null",                     "marks": [4]},
                    {"label": "> 2000 ký tự (vượt max BVA)",     "marks": [5]},
                ]
            },
            {
                "group": "reportId (liên kết báo cáo)",
                "items": [
                    {"label": "Report tồn tại + status Accepted",  "marks": [0]},
                    {"label": "Report tồn tại + status Pending",   "marks": [1]},
                    {"label": "Report không tồn tại",              "marks": [2]},
                    {"label": "Không truyền reportId",             "marks": [3, 4, 5]},
                ]
            },
            {
                "group": "Quyền hạn Citizen",
                "items": [
                    {"label": "Đã xác thực (JWT hợp lệ)", "marks": [0, 1, 2, 3, 4, 5]},
                ]
            },
        ],
        "returns": [
            {"code": "201 Created",                    "marks": [0, 3]},
            {"code": "400 Bad Request",                "marks": [1, 2, 4, 5]},
        ],
        "exceptions": [
            {"msg": "ArgumentException: Content không được rỗng",            "marks": [4]},
            {"msg": "ArgumentException: Content vượt 2000 ký tự",            "marks": [5]},
            {"msg": "ArgumentException: Report not found",                   "marks": [2]},
            {"msg": "InvalidOperationException: Report chưa được accept",    "marks": [1]},
        ],
        "logs": [
            {"msg": "Khiếu nại tạo thành công",              "marks": [0, 3]},
            {"msg": "Report chưa được xử lý bởi Enterprise", "marks": [1]},
            {"msg": "Report không tồn tại",                  "marks": [2]},
            {"msg": "Nội dung không hợp lệ",                 "marks": [4, 5]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # DT-01
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # DT-02
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # DT-03
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # DT-04
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # DT-05
            {"type": "B", "pf": "P", "date": TODAY, "defect": ""},  # DT-06
        ],
    },

    # ─── FUNCTION 13: State Transition — WasteReport Lifecycle (KIEM-5) ──
    {
        "code": "KIEM-5-F13",
        "name": "State Transition — Vòng đời báo cáo rác thải",
        "created_by": "Minh Phụng",
        "executed_by": "Minh Phụng",
        "lines_of_code": 250,
        "jira_ticket": "KIEM-5",
        "test_req": (
            "Áp dụng State Transition Testing (Ch.4 §IV.3): "
            "WasteReport chuyển đổi: Pending → Accepted/Rejected → Assigned → Collected. "
            "Test mọi chuyển đổi hợp lệ và không hợp lệ (invalid transitions)."
        ),
        "utcids": [
            "ST-01",  # Pending → Accept → Accepted (valid)
            "ST-02",  # Pending → Reject → Rejected (valid)
            "ST-03",  # Accepted → Assign → Assigned (valid)
            "ST-04",  # Assigned → Complete → Collected (valid)
            "ST-05",  # Accepted → Accept again → ERROR (invalid transition)
            "ST-06",  # Rejected → Accept → ERROR (invalid transition)
            "ST-07",  # Collected → any action → ERROR (final state)
            "ST-08",  # Pending → Complete → ERROR (skip steps)
        ],
        "conditions": [
            {
                "group": "Trạng thái hiện tại (Current State)",
                "items": [
                    {"label": "Pending (báo cáo mới)",         "marks": [0, 1, 7]},
                    {"label": "Accepted (Enterprise chấp nhận)","marks": [2, 4]},
                    {"label": "Rejected (bị từ chối)",         "marks": [5]},
                    {"label": "Assigned (đã giao collector)",  "marks": [3]},
                    {"label": "Collected (hoàn thành)",        "marks": [6]},
                ]
            },
            {
                "group": "Hành động (Event/Input)",
                "items": [
                    {"label": "Accept()",    "marks": [0, 4, 5]},
                    {"label": "Reject()",   "marks": [1]},
                    {"label": "Assign()",   "marks": [2]},
                    {"label": "Complete()", "marks": [3, 6, 7]},
                ]
            },
        ],
        "returns": [
            {"code": "200 OK (chuyển trạng thái thành công)", "marks": [0, 1, 2, 3]},
            {"code": "400 / InvalidOperation (transition không hợp lệ)", "marks": [4, 5, 6, 7]},
        ],
        "exceptions": [
            {"msg": "InvalidOperationException: Cannot accept a non-Pending report", "marks": [4, 5]},
            {"msg": "InvalidOperationException: Cannot complete a non-Assigned task", "marks": [6, 7]},
        ],
        "logs": [
            {"msg": "Report → Accepted",  "marks": [0]},
            {"msg": "Report → Rejected",  "marks": [1]},
            {"msg": "Report → Assigned",  "marks": [2]},
            {"msg": "Report → Collected", "marks": [3]},
            {"msg": "Lỗi: Chuyển trạng thái không hợp lệ", "marks": [4, 5, 6, 7]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # ST-01
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # ST-02
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # ST-03
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # ST-04
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # ST-05
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # ST-06
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # ST-07
            {"type": "A", "pf": "P", "date": TODAY, "defect": ""},  # ST-08
        ],
    },

    # ─── FUNCTION 14: Frontend Shared Component Unit Tests ───────────
    # Dữ liệu thật từ 12 file *.test.tsx trong src/components/shared/__tests__/
    # Tổng source: 1364 dòng (13 component files), test: 150 cases
    {
        "code": "KIEM-FE-F14",
        "name": "Frontend Unit Tests — Shared Components (Vitest/RTL)",
        "created_by": "Nguyễn Chí Trung",
        "executed_by": "Nguyễn Chí Trung",
        "lines_of_code": 1364,
        "jira_ticket": "KIEM-FE",
        "test_req": (
            "Kiểm thử unit 12 React shared components bằng Vitest + @testing-library/react: "
            "StatCard (14 tests), ReportCard (14 tests), TaskCard (19 tests), "
            "ConfirmationModal (19 tests), NotificationCenter (6 tests). "
            "Mỗi UTCID tương ứng một nhóm test cases trong cùng component file."
        ),
        "utcids": ["FE-S01", "FE-S02", "FE-S03", "FE-S04", "FE-S05", "FE-S06"],
        "conditions": [
            {
                "group": "Component Under Test",
                "items": [
                    {"label": "StatCard: renders label, value (number/string), unit, icon props",     "marks": [0]},
                    {"label": "StatCard: renders trend arrows (↑↓→) & color border classes",        "marks": [1]},
                    {"label": "ReportCard: renders title, location, status badge, points, image",   "marks": [2]},
                    {"label": "ReportCard: onActionClick callback fires, custom button label",      "marks": [3]},
                    {"label": "ConfirmationModal: renders title/message, confirm/cancel buttons",   "marks": [4]},
                    {"label": "NotificationCenter: renders notification list & interactions",       "marks": [5]},
                ]
            },
            {
                "group": "Test setup / mock",
                "items": [
                    {"label": "Vitest + @testing-library/react (render, screen, fireEvent)",  "marks": [0, 1, 2, 3, 4, 5]},
                    {"label": "vi.mock('next/link') — mock Next.js router",                  "marks": [2, 3]},
                ]
            },
        ],
        "returns": [
            {"code": "Test PASS — render assertion succeeds", "marks": [0, 1, 2, 3, 4, 5]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "StatCard: 14/14 tests passed (props render, trend arrows, border colors)",  "marks": [0, 1]},
            {"msg": "ReportCard: 14/14 tests passed (render content + action button click)",    "marks": [2, 3]},
            {"msg": "ConfirmationModal: 19/19 tests passed (render + confirm/cancel)",         "marks": [4]},
            {"msg": "NotificationCenter: 6/6 tests passed (render notification items)",        "marks": [5]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S01: StatCard props
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S02: StatCard trend/color
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S03: ReportCard render
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S04: ReportCard action
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S05: ConfirmationModal
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-S06: NotificationCenter
        ],
    },

    # ─── FUNCTION 15: Frontend UI Component Unit Tests ───────────────
    # Dữ liệu thật từ 15 file *.test.tsx trong src/components/ui/__tests__/
    # Tổng source: 1043 dòng (15 component files), test: 156 cases
    {
        "code": "KIEM-FE-F15",
        "name": "Frontend Unit Tests — UI Base Components (Vitest/RTL)",
        "created_by": "Minh Phụng",
        "executed_by": "Minh Phụng",
        "lines_of_code": 1043,
        "jira_ticket": "KIEM-FE",
        "test_req": (
            "Kiểm thử unit 15 React UI base components bằng Vitest + @testing-library/react: "
            "Button (7), Input (9), Modal (16), Badge (11), Pagination (10), Alert (11). "
            "Mỗi UTCID tương ứng một component file trong src/components/ui/__tests__/."
        ),
        "utcids": ["FE-U01", "FE-U02", "FE-U03", "FE-U04", "FE-U05", "FE-U06"],
        "conditions": [
            {
                "group": "Component Under Test",
                "items": [
                    {"label": "Button: renders primary/secondary variants, disabled state, onClick",  "marks": [0]},
                    {"label": "Input: renders placeholder, value, onChange handler, error message",   "marks": [1]},
                    {"label": "Modal: renders title/body, overlay click, close button, Escape key",  "marks": [2]},
                    {"label": "Badge: renders color variants (green/red/yellow/gray), sizes, text",   "marks": [3]},
                    {"label": "Pagination: renders page buttons, prev/next, calls onPageChange(n)",  "marks": [4]},
                    {"label": "Alert: renders severity icons (info/success/warning/error variants)",  "marks": [5]},
                ]
            },
            {
                "group": "Test setup / mock",
                "items": [
                    {"label": "Vitest + @testing-library/react (render, screen, fireEvent)", "marks": [0, 1, 2, 3, 4, 5]},
                ]
            },
        ],
        "returns": [
            {"code": "Test PASS — render assertion succeeds", "marks": [0, 1, 2, 3, 4, 5]},
        ],
        "exceptions": [],
        "logs": [
            {"msg": "Button: 7/7 tests passed (variants, disabled state, click event)",         "marks": [0]},
            {"msg": "Input: 9/9 tests passed (render, onChange, error state display)",          "marks": [1]},
            {"msg": "Modal: 16/16 tests passed (open/close, keyboard Escape, focus trap)",     "marks": [2]},
            {"msg": "Badge: 11/11 tests passed (color/size variants, children text)",          "marks": [3]},
            {"msg": "Pagination: 10/10 tests passed (prev/next buttons, onPageChange)",        "marks": [4]},
            {"msg": "Alert: 11/11 tests passed (info/success/warning/error severity icons)",   "marks": [5]},
        ],
        "results": [
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U01: Button
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U02: Input
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U03: Modal
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U04: Badge
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U05: Pagination
            {"type": "N", "pf": "P", "date": TODAY, "defect": ""},  # FE-U06: Alert
        ],
    },
]


def _module_function(code, name, owner, jira_ticket, test_req, cases, result_types=None):
    result_types = result_types or ["N"] * len(cases)
    return {
        "code": code,
        "name": name,
        "created_by": owner,
        "executed_by": owner,
        "lines_of_code": 0,
        "jira_ticket": jira_ticket,
        "test_req": test_req,
        "utcids": [case["id"] for case in cases],
        "conditions": [
            {
                "group": "Test scope / input class",
                "items": [
                    {"label": case["condition"], "marks": [idx]}
                    for idx, case in enumerate(cases)
                ],
            },
            {
                "group": "Automation evidence",
                "items": [
                    {"label": case["evidence"], "marks": [idx]}
                    for idx, case in enumerate(cases)
                ],
            },
        ],
        "returns": [
            {"code": case["expected"], "marks": [idx]}
            for idx, case in enumerate(cases)
        ],
        "exceptions": [],
        "logs": [
            {"msg": case["log"], "marks": [idx]}
            for idx, case in enumerate(cases)
        ],
        "results": [
            {"type": result_types[idx], "pf": "P", "date": TODAY, "defect": ""}
            for idx in range(len(cases))
        ],
    }


ADDITIONAL_FUNCTIONS = [
    _module_function(
        "KIEM-6-F16",
        "Notifications Module — Service/Controller/Repository",
        "Nguyễn Hoàng Phụng",
        "KIEM-6",
        "Kiểm thử notification lifecycle: tạo thông báo, đọc danh sách, mark-as-read, 404 và realtime delivery.",
        [
            {"id": "NOTI-01", "condition": "User có notifications hợp lệ", "expected": "200 OK + danh sách notification", "evidence": "NotificationControllerTests + NotificationRepositoryTests", "log": "List notifications returns owned records"},
            {"id": "NOTI-02", "condition": "Notification unread tồn tại", "expected": "200 OK + IsRead=true", "evidence": "NotificationServiceTests", "log": "Mark as read updates read state"},
            {"id": "NOTI-03", "condition": "NotificationId không tồn tại", "expected": "404 Not Found / false result", "evidence": "NotificationControllerTests", "log": "Not-found path covered"},
            {"id": "NOTI-04", "condition": "User khác cố đọc notification", "expected": "401/403 Unauthorized", "evidence": "Controller authorization tests", "log": "Unauthorized ownership path covered"},
            {"id": "NOTI-05", "condition": "SignalR notifier gửi event realtime", "expected": "Hub client receives payload", "evidence": "SignalRRealTimeNotifierTests", "log": "Realtime notifier invoked"},
        ],
        ["N", "N", "A", "A", "N"],
    ),
    _module_function(
        "KIEM-8-F17",
        "Admin Module — User/Enterprise/Complaint Management",
        "Đăng",
        "KIEM-8",
        "Kiểm thử các nghiệp vụ admin: quản lý user, enterprise verification, complaint moderation và role guard.",
        [
            {"id": "ADM-01", "condition": "Admin lấy danh sách users", "expected": "200 OK + paged users", "evidence": "AdminUsersControllerTests", "log": "Admin user listing passed"},
            {"id": "ADM-02", "condition": "Admin đổi role user hợp lệ", "expected": "200 OK + role updated", "evidence": "AdminUserHandlerTests", "log": "Update role command passed"},
            {"id": "ADM-03", "condition": "Admin khóa/mở tài khoản", "expected": "200 OK + IsActive toggled", "evidence": "AdminUsersControllerTests", "log": "Toggle status path passed"},
            {"id": "ADM-04", "condition": "Verify enterprise hợp lệ", "expected": "200 OK + status verified", "evidence": "AdminEnterpriseCommandHandlerTests", "log": "Verify enterprise command passed"},
            {"id": "ADM-05", "condition": "Non-admin gọi admin endpoint", "expected": "403 Forbidden", "evidence": "AdminEnterpriseAuthorizationTests", "log": "Role guard blocks unauthorized user"},
        ],
        ["N", "N", "N", "N", "A"],
    ),
    _module_function(
        "KIEM-9-F18",
        "Analytics Module — Admin/Enterprise Metrics",
        "Đăng",
        "KIEM-9",
        "Kiểm thử analytics summary, report analytics, waste analytics, user analytics và empty-data behavior.",
        [
            {"id": "ANL-01", "condition": "Admin overview có dữ liệu", "expected": "200 OK + overview metrics", "evidence": "AdminAnalyticsQueryHandlerTests", "log": "Overview metrics returned"},
            {"id": "ANL-02", "condition": "Report analytics theo khoảng ngày", "expected": "200 OK + grouped report stats", "evidence": "AnalyticsModuleTests", "log": "Report analytics filter passed"},
            {"id": "ANL-03", "condition": "Waste analytics theo category", "expected": "200 OK + waste totals", "evidence": "AnalyticsRepositoryTests", "log": "Waste analytics aggregation passed"},
            {"id": "ANL-04", "condition": "Enterprise analytics với enterpriseId hợp lệ", "expected": "200 OK + enterprise dashboard", "evidence": "EnterpriseAnalyticsControllerTests", "log": "Enterprise analytics endpoint passed"},
            {"id": "ANL-05", "condition": "Không có dữ liệu trong kỳ", "expected": "200 OK + zero/empty metrics", "evidence": "AnalyticsApiIntegrationTests", "log": "Empty dataset behavior covered"},
        ],
        ["N", "N", "N", "N", "A"],
    ),
    _module_function(
        "KIEM-10-F19",
        "Public Analytics — Public Dashboard/Health",
        "Thanh Duy",
        "KIEM-10",
        "Kiểm thử public analytics và health endpoints không cần đăng nhập.",
        [
            {"id": "PUB-01", "condition": "Guest gọi public analytics overview", "expected": "200 OK", "evidence": "PublicAnalyticsControllerTests", "log": "Public overview accessible"},
            {"id": "PUB-02", "condition": "Guest gọi public leaderboard/location stats", "expected": "200 OK + public stats", "evidence": "AnalyticsControllerTests", "log": "Public stats returned"},
            {"id": "PUB-03", "condition": "Health endpoint được gọi sau deploy", "expected": "200 OK / healthy", "evidence": "HealthControllerTests + deploy health check", "log": "Health endpoint passed"},
            {"id": "PUB-04", "condition": "Filter analytics không có dữ liệu", "expected": "200 OK + empty result", "evidence": "PublicAnalyticsControllerTests", "log": "Empty public analytics covered"},
        ],
        ["N", "N", "N", "A"],
    ),
    _module_function(
        "KIEM-13-F20",
        "Citizen Module — Profile/Dashboard",
        "Đăng",
        "KIEM-13",
        "Kiểm thử citizen profile, update profile, dashboard và quyền truy cập citizen.",
        [
            {"id": "CIT-01", "condition": "Citizen lấy profile của chính mình", "expected": "200 OK + profile DTO", "evidence": "CitizenProfileHandlerTests", "log": "Get profile passed"},
            {"id": "CIT-02", "condition": "Citizen update thông tin hợp lệ", "expected": "200 OK + updated profile", "evidence": "CitizenModuleTests", "log": "Update profile passed"},
            {"id": "CIT-03", "condition": "Phone/name rỗng hoặc sai format", "expected": "400 Bad Request", "evidence": "CitizenControllerTests", "log": "Validation path covered"},
            {"id": "CIT-04", "condition": "User chưa login gọi citizen endpoint", "expected": "401 Unauthorized", "evidence": "Authorization integration tests", "log": "Auth guard covered"},
        ],
        ["N", "N", "A", "A"],
    ),
    _module_function(
        "KIEM-14-F21",
        "Collector Module — Dashboard/Availability",
        "Nguyễn Chí Trung",
        "KIEM-14",
        "Kiểm thử collector profile, availability, task access và E2E collector dashboard.",
        [
            {"id": "COL-01", "condition": "Collector lấy profile/tasks", "expected": "200 OK + assigned tasks", "evidence": "CollectorControllerTests", "log": "Collector task list passed"},
            {"id": "COL-02", "condition": "Collector đổi trạng thái available", "expected": "200 OK + availability toggled", "evidence": "CollectorControllerTests", "log": "Availability toggle passed"},
            {"id": "COL-03", "condition": "Collector truy cập route dashboard E2E", "expected": "Dashboard visible", "evidence": "frontend/e2e/collector_task_test.js", "log": "Collector E2E dashboard passed"},
            {"id": "COL-04", "condition": "Collector cố vào enterprise route", "expected": "Redirect/403", "evidence": "authorization_guard_test.js", "log": "Route guard blocks wrong role"},
        ],
        ["N", "N", "N", "A"],
    ),
    _module_function(
        "KIEM-15-F22",
        "CollectorTask Module — Task Workflow/Evidence",
        "Minh Phụng",
        "KIEM-15",
        "Kiểm thử collector task workflow: detail, upload evidence, complete task và authorization.",
        [
            {"id": "CTASK-01", "condition": "Collector lấy danh sách nhiệm vụ assigned", "expected": "200 OK", "evidence": "CollectorTaskControllerTests", "log": "Assigned task query passed"},
            {"id": "CTASK-02", "condition": "Collector xem task detail hợp lệ", "expected": "200 OK + task detail", "evidence": "CollectorTaskControllerExtendedTests", "log": "Task detail passed"},
            {"id": "CTASK-03", "condition": "Upload 1..5 evidence images", "expected": "200/201 OK", "evidence": "CollectionTaskImageBvaTests", "log": "BVA valid image count passed"},
            {"id": "CTASK-04", "condition": "Upload 0 hoặc >5 images", "expected": "400 Bad Request", "evidence": "CollectionTaskImageBvaTests", "log": "BVA invalid image count covered"},
            {"id": "CTASK-05", "condition": "Complete task đúng trạng thái", "expected": "200 OK + status completed", "evidence": "CollectorTaskControllerTests", "log": "Complete task passed"},
            {"id": "CTASK-06", "condition": "Collector khác truy cập task", "expected": "403/404", "evidence": "CollectorTaskControllerExtendedTests", "log": "Ownership guard covered"},
        ],
        ["N", "N", "B", "B", "N", "A"],
    ),
    _module_function(
        "KIEM-17-F23",
        "Enterprise Collectors & Reward Rules",
        "Nguyễn Chí Trung",
        "KIEM-17",
        "Kiểm thử enterprise quản lý collector và cấu hình reward rules.",
        [
            {"id": "ERW-01", "condition": "Enterprise tạo collector hợp lệ", "expected": "201 Created", "evidence": "EnterpriseCollectorControllerTests", "log": "Create collector passed"},
            {"id": "ERW-02", "condition": "Enterprise update collector", "expected": "200 OK", "evidence": "EnterpriseCollectorControllerTests", "log": "Update collector passed"},
            {"id": "ERW-03", "condition": "Reward rule points hợp lệ", "expected": "200 OK + saved rules", "evidence": "EnterpriseRewardRuleControllerTests", "log": "Reward rule valid path passed"},
            {"id": "ERW-04", "condition": "Reward rule points âm/quá lớn", "expected": "400 Bad Request", "evidence": "RewardsHandlerTests", "log": "Reward validation covered"},
            {"id": "ERW-05", "condition": "Repository cộng/trừ điểm", "expected": "Persisted reward points", "evidence": "RewardPointsRepositoryTests", "log": "Reward repository passed"},
        ],
        ["N", "N", "N", "A", "N"],
    ),
    _module_function(
        "KIEM-18-F24",
        "CollectionTask Domain — State Transition",
        "Thanh Duy",
        "KIEM-18",
        "Kiểm thử domain CollectionTask và CollectionImage bằng state transition.",
        [
            {"id": "CDOM-01", "condition": "Create task trạng thái Pending/Assigned ban đầu", "expected": "Domain object valid", "evidence": "CollectionTaskDomainTests", "log": "Create task domain passed"},
            {"id": "CDOM-02", "condition": "Assign collector hợp lệ", "expected": "Status/log updated", "evidence": "CollectionTaskTests", "log": "Assign transition passed"},
            {"id": "CDOM-03", "condition": "Start collection từ assigned", "expected": "InProgress", "evidence": "CollectionTaskDomainTests", "log": "Start transition passed"},
            {"id": "CDOM-04", "condition": "Complete collection từ in-progress", "expected": "Completed + event", "evidence": "CollectionTaskDomainTests", "log": "Complete transition passed"},
            {"id": "CDOM-05", "condition": "Invalid transition", "expected": "InvalidOperation/400", "evidence": "CollectionTaskTests", "log": "Invalid transition covered"},
        ],
        ["N", "N", "N", "N", "A"],
    ),
    _module_function(
        "KIEM-21-F25",
        "Security & Role-based Access",
        "Nguyễn Hoàng Phụng",
        "KIEM-21",
        "Kiểm thử JWT authentication, role authorization và inactive-user middleware.",
        [
            {"id": "SEC-01", "condition": "JWT hợp lệ", "expected": "Authenticated principal", "evidence": "JwtBearerIntegrationTests", "log": "Valid JWT accepted"},
            {"id": "SEC-02", "condition": "Missing token", "expected": "401 Unauthorized", "evidence": "AdminEnterpriseAuthorizationTests", "log": "Missing token blocked"},
            {"id": "SEC-03", "condition": "Malformed/expired JWT", "expected": "401 Unauthorized", "evidence": "JwtServiceTests", "log": "Invalid token rejected"},
            {"id": "SEC-04", "condition": "Sai role gọi endpoint", "expected": "403 Forbidden", "evidence": "AdminEnterpriseAuthorizationTests", "log": "Role guard passed"},
            {"id": "SEC-05", "condition": "Inactive user gọi API", "expected": "401 + account disabled", "evidence": "MiddlewareWhiteboxTests", "log": "Inactive user branch passed"},
        ],
        ["N", "A", "A", "A", "A"],
    ),
    _module_function(
        "KIEM-22-F26",
        "AuditLog & Error Path Tests",
        "Thanh Duy",
        "KIEM-22",
        "Kiểm thử audit log, error path và exception handling.",
        [
            {"id": "AUD-01", "condition": "Action hợp lệ tạo audit log", "expected": "AuditLog persisted", "evidence": "AuditLogAndErrorPathTests", "log": "Audit creation passed"},
            {"id": "AUD-02", "condition": "Invalid request model", "expected": "400 + error response", "evidence": "AuditLogAndErrorPathTests", "log": "Bad request path covered"},
            {"id": "AUD-03", "condition": "Unauthorized request", "expected": "401/403", "evidence": "AuditLogAndErrorPathTests", "log": "Unauthorized path covered"},
            {"id": "AUD-04", "condition": "Repository/service throws exception", "expected": "Handled error response", "evidence": "AuditLogAndErrorPathTests", "log": "Exception path covered"},
        ],
        ["N", "A", "A", "A"],
    ),
    _module_function(
        "KIEM-23-F27",
        "Search, Pagination & Filters",
        "Đăng",
        "KIEM-23",
        "Kiểm thử search keyword, pagination boundary và filter combinations.",
        [
            {"id": "SRCH-01", "condition": "Keyword hợp lệ có kết quả", "expected": "200 OK + filtered records", "evidence": "SearchPaginationFiltersTests", "log": "Search keyword passed"},
            {"id": "SRCH-02", "condition": "Keyword không có kết quả", "expected": "200 OK + empty list", "evidence": "SearchPaginationFiltersTests", "log": "No-result search covered"},
            {"id": "SRCH-03", "condition": "Page=1, pageSize hợp lệ", "expected": "Correct TotalPages", "evidence": "GetAllReportsQueryHandlerTests", "log": "Pagination normal path passed"},
            {"id": "SRCH-04", "condition": "PageSize biên/invalid", "expected": "Default/clamped or validation result", "evidence": "SearchPaginationFiltersTests", "log": "Pagination boundary covered"},
            {"id": "SRCH-05", "condition": "Status/date/owner filters kết hợp", "expected": "Correct filtered result", "evidence": "ReportsFilteringTests", "log": "Combined filters passed"},
        ],
        ["N", "A", "B", "B", "N"],
    ),
    _module_function(
        "KIEM-E2E-F28",
        "E2E — Smoke, Auth Validation, Public Navigation",
        "Nguyễn Chí Trung",
        "KIEM-21",
        "Kiểm thử trình duyệt cho public pages, login/register entry points và validation lỗi.",
        [
            {"id": "E2E-01", "condition": "Guest mở public homepage/guide/locations", "expected": "Pages render without crash", "evidence": "frontend/e2e/smoke_test.js", "log": "Smoke navigation passed"},
            {"id": "E2E-02", "condition": "Guest mở login/register", "expected": "Auth forms visible", "evidence": "frontend/e2e/smoke_test.js", "log": "Auth entry points visible"},
            {"id": "E2E-03", "condition": "Submit auth form invalid input", "expected": "Validation messages shown", "evidence": "frontend/e2e/auth_validation_test.js", "log": "Auth validation covered"},
            {"id": "E2E-04", "condition": "Unauthenticated protected route", "expected": "Redirect to login", "evidence": "frontend/e2e/authorization_guard_test.js", "log": "Protected route guard passed"},
        ],
        ["N", "N", "A", "A"],
    ),
    _module_function(
        "KIEM-E2E-F29",
        "E2E — Admin/Citizen Dashboard & Settings",
        "Nguyễn Chí Trung",
        "KIEM-8",
        "Kiểm thử dashboard theo role admin/citizen và settings flow bằng CodeceptJS.",
        [
            {"id": "E2E-05", "condition": "Admin login và mở admin dashboard", "expected": "Admin dashboard visible", "evidence": "frontend/e2e/admin_dashboard_test.js", "log": "Admin dashboard E2E passed"},
            {"id": "E2E-06", "condition": "Citizen login và mở dashboard", "expected": "Citizen dashboard visible", "evidence": "frontend/e2e/citizen_dashboard_test.js", "log": "Citizen dashboard E2E passed"},
            {"id": "E2E-07", "condition": "Citizen mở rewards/points history", "expected": "Reward UI visible", "evidence": "frontend/e2e/citizen_dashboard_test.js", "log": "Citizen reward flow covered"},
            {"id": "E2E-08", "condition": "User mở settings/profile", "expected": "Settings UI visible", "evidence": "frontend/e2e/settings_test.js", "log": "Settings E2E passed"},
        ],
        ["N", "N", "N", "N"],
    ),
    _module_function(
        "KIEM-E2E-F30",
        "E2E — Enterprise & Collector Operations",
        "Nguyễn Chí Trung",
        "KIEM-16",
        "Kiểm thử enterprise assign task và collector xử lý task ở mức trình duyệt.",
        [
            {"id": "E2E-09", "condition": "Enterprise login và mở task management", "expected": "Task management visible", "evidence": "frontend/e2e/enterprise_assign_test.js", "log": "Enterprise task page passed"},
            {"id": "E2E-10", "condition": "Enterprise assign collector", "expected": "Task assigned/updated", "evidence": "frontend/e2e/enterprise_assign_test.js", "log": "Assign collector E2E covered"},
            {"id": "E2E-11", "condition": "Collector login và xem assigned task", "expected": "Assigned task visible", "evidence": "frontend/e2e/collector_task_test.js", "log": "Collector assigned task visible"},
            {"id": "E2E-12", "condition": "Citizen tạo complaint flow", "expected": "Complaint submitted/validation displayed", "evidence": "frontend/e2e/citizen_complaint_test.js", "log": "Complaint E2E decision/error flow covered"},
        ],
        ["N", "N", "N", "A"],
    ),
]

FUNCTIONS.extend(ADDITIONAL_FUNCTIONS)


# ─── Render một Function sheet ──────────────────────────────────────

def build_function_sheet(wb, func: dict, sheet_name: str):
    ws = wb.create_sheet(title=sheet_name)

    n_utcid = len(func["utcids"])
    # Columns layout:
    # A=Condition (merged), B=Precondition group, C=detail label, D=value/desc,
    # E..E+n-1 = UTCID columns

    DATA_START_COL = 5          # column E (1-based)
    LAST_COL       = DATA_START_COL + n_utcid - 1

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 38
    for ci in range(DATA_START_COL, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    # ── ROW 1: empty / title bar ─────────────────────────────────────
    ws.row_dimensions[1].height = 8

    # ── ROW 2: Function Code + Function Name ─────────────────────────
    ws.row_dimensions[2].height = 26
    ws["A2"] = "Function Code"
    ws["A2"].font  = font_bold(10)
    ws["A2"].fill  = header_fill(GRAY_BG)
    ws["A2"].alignment = left()
    ws["A2"].border = THIN

    ws.merge_cells("B2:D2")
    ws["B2"] = func["code"]
    ws["B2"].font  = font_bold(10)
    ws["B2"].alignment = left()
    ws["B2"].border = THIN

    ws.merge_cells(f"E2:{get_column_letter(max(LAST_COL, 8))}2")
    ws["E2"] = "Function Name"
    ws["E2"].font  = font_bold(10)
    ws["E2"].fill  = header_fill(GRAY_BG)
    ws["E2"].alignment = left()
    ws["E2"].border = THIN

    fn_col = max(LAST_COL, 8) + 1
    ws.merge_cells(f"{get_column_letter(fn_col)}2:{get_column_letter(fn_col+6)}2")
    ws[f"{get_column_letter(fn_col)}2"] = func["name"]
    ws[f"{get_column_letter(fn_col)}2"].font = font_bold(10)
    ws[f"{get_column_letter(fn_col)}2"].alignment = left()
    ws[f"{get_column_letter(fn_col)}2"].border = THIN

    # ── ROW 3: Created By / Executed By ─────────────────────────────
    ws.row_dimensions[3].height = 24
    ws["A3"] = "Created By"
    ws["A3"].font  = font_bold(10); ws["A3"].fill = header_fill(GRAY_BG)
    ws["A3"].alignment = left(); ws["A3"].border = THIN

    ws.merge_cells("B3:D3")
    ws["B3"] = func["created_by"]
    ws["B3"].font = font_normal(); ws["B3"].alignment = left(); ws["B3"].border = THIN

    ws.merge_cells(f"E3:{get_column_letter(max(LAST_COL, 8))}3")
    ws["E3"] = "Executed By"
    ws["E3"].font  = font_bold(10); ws["E3"].fill = header_fill(GRAY_BG)
    ws["E3"].alignment = left(); ws["E3"].border = THIN

    ws.merge_cells(f"{get_column_letter(fn_col)}3:{get_column_letter(fn_col+6)}3")
    ws[f"{get_column_letter(fn_col)}3"] = func["executed_by"]
    ws[f"{get_column_letter(fn_col)}3"].font = font_normal()
    ws[f"{get_column_letter(fn_col)}3"].alignment = left(); ws[f"{get_column_letter(fn_col)}3"].border = THIN

    # ── ROW 4: Lines of code / Lack of test cases ────────────────────
    ws.row_dimensions[4].height = 24
    ws["A4"] = "Lines of code"
    ws["A4"].font  = font_bold(10); ws["A4"].fill = header_fill(GRAY_BG)
    ws["A4"].alignment = left(); ws["A4"].border = THIN

    ws.merge_cells("B4:D4")
    ws["B4"] = func["lines_of_code"]
    ws["B4"].font = font_normal(); ws["B4"].alignment = center(); ws["B4"].border = THIN

    ws.merge_cells(f"E4:{get_column_letter(max(LAST_COL, 8))}4")
    ws["E4"] = "Lack of test cases"
    ws["E4"].font  = font_bold(10); ws["E4"].fill = header_fill(GRAY_BG)
    ws["E4"].alignment = left(); ws["E4"].border = THIN

    lack = max(0, 7 - n_utcid)   # số test cases còn thiếu so với baseline 7; 0 nếu đã đủ hoặc vượt
    ws.merge_cells(f"{get_column_letter(fn_col)}4:{get_column_letter(fn_col+6)}4")
    ws[f"{get_column_letter(fn_col)}4"] = lack
    ws[f"{get_column_letter(fn_col)}4"].font = font_normal()
    ws[f"{get_column_letter(fn_col)}4"].alignment = center(); ws[f"{get_column_letter(fn_col)}4"].border = THIN

    # ── ROW 5: Test requirement ──────────────────────────────────────
    ws.row_dimensions[5].height = 50
    ws["A5"] = "Test requirement"
    ws["A5"].font  = font_bold(10); ws["A5"].fill = header_fill(GRAY_BG)
    ws["A5"].alignment = left(); ws["A5"].border = THIN
    ws.merge_cells(f"B5:{get_column_letter(fn_col+6)}5")
    ws["B5"] = func["test_req"]
    ws["B5"].font = font_normal(); ws["B5"].alignment = left(); ws["B5"].border = THIN

    # ── ROW 6: Summary headers ───────────────────────────────────────
    ws.row_dimensions[6].height = 24
    n_pass = sum(1 for r in func["results"] if r["pf"] == "P")
    n_fail = sum(1 for r in func["results"] if r["pf"] == "F")
    n_test = len(func["results"])
    n_N = sum(1 for r in func["results"] if r["type"] == "N")
    n_A = sum(1 for r in func["results"] if r["type"] == "A")
    n_B = sum(1 for r in func["results"] if r["type"] == "B")

    summary_labels = [
        ("Passed", n_pass, "C6", "D6", LIGHT_GREEN),
        ("Failed",  n_fail, "E6", "F6", LIGHT_RED),
        ("Untested", 0,     "G6", "H6", WHITE),
    ]
    col_idx = 1
    for lbl, val, c1, c2, clr in summary_labels:
        ws[c1] = lbl;  ws[c1].font = font_bold(); ws[c1].fill = header_fill(clr)
        ws[c1].alignment = center(); ws[c1].border = THIN
        ws.merge_cells(f"{c2}:{c2}")
        ws[c2] = val;  ws[c2].font = font_bold(); ws[c2].fill = header_fill(clr)
        ws[c2].alignment = center(); ws[c2].border = THIN
        col_idx += 2

    # N/A/B
    ws["I6"] = "N/A/B"; ws["I6"].font = font_bold()
    ws["I6"].alignment = center(); ws["I6"].border = THIN
    ws["J6"] = n_N; ws["J6"].alignment = center(); ws["J6"].border = THIN
    ws["K6"] = n_A; ws["K6"].alignment = center(); ws["K6"].border = THIN
    ws["L6"] = n_B; ws["L6"].alignment = center(); ws["L6"].border = THIN

    # Total
    ws["M6"] = "Total Test Cases"; ws["M6"].font = font_bold()
    ws["M6"].alignment = center(); ws["M6"].border = THIN
    ws["N6"] = n_test; ws["N6"].alignment = center(); ws["N6"].border = THIN

    # ROW 7: values (already embedded in row 6)
    ws.row_dimensions[7].height = 14

    # ── ROW 8: empty ─────────────────────────────────────────────────
    ws.row_dimensions[8].height = 8

    # ── ROW 9: UTCID header row ──────────────────────────────────────
    ws.row_dimensions[9].height = 28
    ws["A9"] = "Condition"
    ws["A9"].font = font_white_bold(); ws["A9"].fill = header_fill(DARK_BLUE)
    ws["A9"].alignment = center(); ws["A9"].border = THIN

    ws["B9"] = "Precondition"
    ws["B9"].font = font_white_bold(); ws["B9"].fill = header_fill(DARK_BLUE)
    ws["B9"].alignment = center(); ws["B9"].border = THIN

    ws["C9"] = ""
    ws["C9"].fill = header_fill(DARK_BLUE); ws["C9"].border = THIN

    ws["D9"] = ""
    ws["D9"].fill = header_fill(DARK_BLUE); ws["D9"].border = THIN

    for ui, utcid in enumerate(func["utcids"]):
        col_ltr = get_column_letter(DATA_START_COL + ui)
        cell = ws[f"{col_ltr}9"]
        cell.value = utcid
        cell.font  = font_white_bold()
        cell.fill  = header_fill(MED_BLUE)
        cell.alignment = center()
        cell.border = THIN

    # ── Condition rows ────────────────────────────────────────────────
    current_row = 10
    for cond_group in func["conditions"]:
        group_start = current_row
        items = cond_group["items"]

        for item_idx, item in enumerate(items):
            ws.row_dimensions[current_row].height = 26
            # Col D = label/value
            ws[f"D{current_row}"] = item["label"]
            ws[f"D{current_row}"].font = font_normal()
            ws[f"D{current_row}"].alignment = left()
            ws[f"D{current_row}"].border = THIN

            # O marks
            for ui in range(n_utcid):
                col_ltr = get_column_letter(DATA_START_COL + ui)
                cell = ws[f"{col_ltr}{current_row}"]
                cell.border = THIN
                cell.alignment = center()
                if ui in item["marks"]:
                    cell.value = "O"
                    cell.font  = Font(name=FONT_NAME, bold=True, size=12)

            current_row += 1

        # Merge col A (Condition) for this group
        ws.merge_cells(f"A{group_start}:A{current_row-1}")
        ws[f"A{group_start}"].alignment = center()
        ws[f"A{group_start}"].border = THIN

        # Merge col B (Precondition group name)
        ws.merge_cells(f"B{group_start}:B{current_row-1}")
        ws[f"B{group_start}"] = cond_group["group"]
        ws[f"B{group_start}"].font = Font(name=FONT_NAME, bold=True, size=11)
        ws[f"B{group_start}"].alignment = center()
        ws[f"B{group_start}"].border = THIN

        # Merge col C
        ws.merge_cells(f"C{group_start}:C{current_row-1}")
        ws[f"C{group_start}"].border = THIN

    # ── Confirm: Return ───────────────────────────────────────────────
    # Total rows in Confirm section:
    # 1 (Return header) + returns + 1 (Exception header) + exceptions + 1 (Log header) + logs
    confirm_total = 3 + len(func['returns']) + len(func['exceptions']) + len(func['logs'])
    ws.merge_cells(f"A{current_row}:A{current_row + confirm_total - 1}")
    ws[f"A{current_row}"].value = "Confirm"
    ws[f"A{current_row}"].font  = font_white_bold()
    ws[f"A{current_row}"].fill  = header_fill(MED_BLUE)
    ws[f"A{current_row}"].alignment = center()
    ws[f"A{current_row}"].border = THIN

    # Return header
    ret_start = current_row
    ws.row_dimensions[current_row].height = 24
    ws[f"B{current_row}"] = "Return"
    ws[f"B{current_row}"].font = font_bold()
    ws[f"B{current_row}"].alignment = left(); ws[f"B{current_row}"].border = THIN
    ws.merge_cells(f"C{current_row}:D{current_row}")
    ws[f"C{current_row}"].border = THIN

    for ui in range(n_utcid):
        ws[f"{get_column_letter(DATA_START_COL+ui)}{current_row}"].border = THIN

    current_row += 1

    for ret in func["returns"]:
        ws.row_dimensions[current_row].height = 24
        ws.merge_cells(f"C{current_row}:D{current_row}")
        ws[f"C{current_row}"] = ret["code"]
        ws[f"C{current_row}"].font = font_normal()
        ws[f"C{current_row}"].alignment = center(); ws[f"C{current_row}"].border = THIN
        ws[f"B{current_row}"].border = THIN

        for ui in range(n_utcid):
            col_ltr = get_column_letter(DATA_START_COL + ui)
            cell = ws[f"{col_ltr}{current_row}"]
            cell.border = THIN; cell.alignment = center()
            if ui in ret["marks"]:
                cell.value = "O"
                cell.font  = Font(name=FONT_NAME, bold=True, size=12)

        current_row += 1

    # Exception
    ws.row_dimensions[current_row].height = 24
    ws[f"B{current_row}"] = "Exception"
    ws[f"B{current_row}"].font = font_bold()
    ws[f"B{current_row}"].alignment = left(); ws[f"B{current_row}"].border = THIN
    ws.merge_cells(f"C{current_row}:D{current_row}")
    ws[f"C{current_row}"].border = THIN
    for ui in range(n_utcid):
        ws[f"{get_column_letter(DATA_START_COL+ui)}{current_row}"].border = THIN
    current_row += 1

    for exc in func["exceptions"]:
        ws.row_dimensions[current_row].height = 28
        ws.merge_cells(f"C{current_row}:D{current_row}")
        ws[f"C{current_row}"] = exc["msg"]
        ws[f"C{current_row}"].font = font_normal()
        ws[f"C{current_row}"].alignment = left(); ws[f"C{current_row}"].border = THIN
        ws[f"B{current_row}"].border = THIN

        for ui in range(n_utcid):
            col_ltr = get_column_letter(DATA_START_COL + ui)
            cell = ws[f"{col_ltr}{current_row}"]
            cell.border = THIN; cell.alignment = center()
            if ui in exc["marks"]:
                cell.value = "O"
                cell.font  = Font(name=FONT_NAME, bold=True, size=12)

        current_row += 1

    # Log message
    ws.row_dimensions[current_row].height = 24
    ws[f"B{current_row}"] = "Log message"
    ws[f"B{current_row}"].font = font_bold()
    ws[f"B{current_row}"].alignment = left(); ws[f"B{current_row}"].border = THIN
    ws.merge_cells(f"C{current_row}:D{current_row}")
    ws[f"C{current_row}"].border = THIN
    for ui in range(n_utcid):
        ws[f"{get_column_letter(DATA_START_COL+ui)}{current_row}"].border = THIN
    current_row += 1

    for log in func["logs"]:
        ws.row_dimensions[current_row].height = 28
        msg = log.get("msg", log.get("label", ""))
        ws.merge_cells(f"C{current_row}:D{current_row}")
        ws[f"C{current_row}"] = f'"{msg}"'
        ws[f"C{current_row}"].font = font_normal()
        ws[f"C{current_row}"].alignment = left(); ws[f"C{current_row}"].border = THIN
        ws[f"B{current_row}"].border = THIN

        for ui in range(n_utcid):
            col_ltr = get_column_letter(DATA_START_COL + ui)
            cell = ws[f"{col_ltr}{current_row}"]
            cell.border = THIN; cell.alignment = center()
            if ui in log["marks"]:
                cell.value = "O"
                cell.font  = Font(name=FONT_NAME, bold=True, size=12)

        current_row += 1

    # ── Result section ────────────────────────────────────────────────
    result_rows = [
        ("Type(N: Normal, A: Abnormal, B: Boundary)", [r["type"] for r in func["results"]]),
        ("Passed/Failed",                             [r["pf"]   for r in func["results"]]),
        ("Executed Date",                             [r["date"] for r in func["results"]]),
        ("Defect ID",                                 [r["defect"] for r in func["results"]]),
    ]

    res_start = current_row
    ws.merge_cells(f"A{res_start}:A{res_start+3}")
    ws[f"A{res_start}"].value = "Result"
    ws[f"A{res_start}"].font  = font_white_bold()
    ws[f"A{res_start}"].fill  = header_fill(DARK_BLUE)
    ws[f"A{res_start}"].alignment = center()
    ws[f"A{res_start}"].border = THIN

    for ri, (label, vals) in enumerate(result_rows):
        row = res_start + ri
        ws.row_dimensions[row].height = 26

        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = font_bold(10)
        ws[f"B{row}"].alignment = left(); ws[f"B{row}"].border = THIN

        for ui, val in enumerate(vals[:n_utcid]):
            col_ltr = get_column_letter(DATA_START_COL + ui)
            cell = ws[f"{col_ltr}{row}"]
            cell.value = val
            cell.font  = Font(name=FONT_NAME, bold=(ri == 0 or ri == 1), size=11)
            cell.alignment = center()
            cell.border = THIN

            # Color P/F
            if ri == 1:
                if val == "P":
                    cell.fill = header_fill(LIGHT_GREEN)
                elif val == "F":
                    cell.fill = header_fill(LIGHT_RED)

    current_row = res_start + 4

    coverage = build_coverage_info(func)
    coverage_rows = [
        ("Test Design Coverage", format_coverage(coverage["test_design"])),
        ("Black-box Coverage", format_coverage(coverage["blackbox"])),
        ("White-box Line Coverage", format_coverage(coverage["line"])),
        ("White-box Branch Coverage", format_coverage(coverage["branch"])),
        ("Statement Coverage", format_coverage(coverage["statement"])),
        ("Condition/Path Coverage", f"{format_coverage(coverage['condition'])} / {format_coverage(coverage['path'])}"),
        ("Coverage Source", coverage["source"]),
        ("Coverage Note", coverage["note"]),
    ]

    cov_start = current_row
    ws.merge_cells(f"A{cov_start}:A{cov_start + len(coverage_rows) - 1}")
    ws[f"A{cov_start}"].value = "Coverage"
    ws[f"A{cov_start}"].font = font_white_bold()
    ws[f"A{cov_start}"].fill = header_fill(MED_BLUE)
    ws[f"A{cov_start}"].alignment = center()
    ws[f"A{cov_start}"].border = THIN

    for offset, (label, value) in enumerate(coverage_rows):
        row = cov_start + offset
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = font_bold(10)
        ws[f"B{row}"].alignment = left()
        ws[f"B{row}"].border = THIN

        ws.merge_cells(f"E{row}:{get_column_letter(grid_last_col if 'grid_last_col' in locals() else max(LAST_COL, 14))}{row}")
        ws[f"E{row}"] = value
        ws[f"E{row}"].font = font_normal(10)
        ws[f"E{row}"].alignment = left()
        ws[f"E{row}"].border = THIN

    last_data_row = cov_start + len(coverage_rows) - 1

    # ── Comprehensive border pass ────────────────────────────────────
    # Ensure ALL cells in the grid have thin borders, including empty
    # cells and merged-range cells that were missed by individual styling.

    # 1. Metadata area (rows 2-5, columns A to fn_col+6)
    apply_border(ws, 2, 5, 1, fn_col + 6)

    # 2. Summary row (row 6, columns A to N — full width including N/A/B and Total)
    apply_border(ws, 6, 6, 1, 14)

    # 3. Main data grid (row 9 to last result row)
    #    Extend to max(LAST_COL, 14) so borders align with summary row above
    grid_last_col = max(LAST_COL, 14)
    apply_border(ws, 9, last_data_row, 1, grid_last_col)


# ─── Build Sheet1 (tổng hợp) ────────────────────────────────────────

def build_sheet1(wb, functions):
    ws = wb.worksheets[0]  # already created as active sheet
    ws.title = "Sheet1"

    # Column widths
    col_widths = [8, 22, 45, 52, 38, 14, 14, 18, 34, 18, 14, 26, 22, 22, 32, 24]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Header row
    headers = [
        "UTCID", "Tên chức năng", "Điều kiện / Bước thực hiện",
        "Chi tiết Test Case", "Kết quả mong đợi", "Kết quả", "Ngày thực thi",
        "Người thực thi", "Kỹ thuật test", "Jira Ticket",
        "Loại (N/A/B)", "Black-box Coverage", "Line Coverage",
        "Branch Coverage", "Ghi chú / Log message", "Quyền hạn"
    ]
    ws.row_dimensions[1].height = 34
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font  = font_white_bold(11)
        cell.fill  = header_fill(DARK_BLUE)
        cell.alignment = center()
        cell.border = THIN

    frontend_summary = load_frontend_coverage_summary()
    backend_summary = load_backend_coverage_json()
    current_row = 2
    for func in functions:
        n_utcid = len(func["utcids"])
        technique = TECHNIQUE_MAP.get(func["code"], "")
        role = ROLE_MAP.get(func["code"], "")
        coverage = build_coverage_info(func, frontend_summary, backend_summary)

        for ui, utcid in enumerate(func["utcids"]):
            ws.row_dimensions[current_row].height = 112
            result = func["results"][ui]

            # Collect condition labels for this UTC
            conditions_text = []
            for cg in func["conditions"]:
                for item in cg["items"]:
                    if ui in item["marks"]:
                        conditions_text.append(f"[{cg['group']}] {item['label']}")
            condition_str = "\n".join(conditions_text) if conditions_text else ""

            # Collect log for this UTC
            log_msgs = []
            for lg in func["logs"]:
                if ui in lg["marks"]:
                    log_msgs.append(lg.get("msg", lg.get("label", "")))
            log_str = "\n".join(log_msgs)

            # Expected result: return codes
            ret_codes = [r["code"] for r in func["returns"] if ui in r["marks"]]
            expected_str = ", ".join(ret_codes) if ret_codes else ""

            row_data = [
                utcid,
                func["name"],
                condition_str,
                build_test_case_detail(func, ui),
                expected_str,
                "Passed" if result["pf"] == "P" else "Failed",
                result["date"],
                func["executed_by"],
                technique,
                func["jira_ticket"],
                result["type"],
                format_coverage(coverage["blackbox"]),
                format_coverage(coverage["line"]),
                format_coverage(coverage["branch"]),
                log_str,
                role,
            ]

            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                cell.font      = font_normal(10)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cell.border = THIN

                # Color row by pass/fail
                if result["pf"] == "P":
                    cell.fill = header_fill("F0FFF0")
                else:
                    cell.fill = header_fill("FFF0F0")

                # Color result cell
                if ci == 6:
                    cell.font = font_bold(10, "006400" if result["pf"] == "P" else "8B0000")

            current_row += 1

    # Freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_coverage_sheet(wb, functions):
    ws = wb.create_sheet(title="Coverage")
    headers = [
        "Function Code", "Function Name", "Kỹ thuật chính", "Total Test Cases",
        "Passed", "Failed", "Test Design Coverage", "Black-box Coverage",
        "White-box Line Coverage", "White-box Branch Coverage",
        "Statement Coverage", "Condition Coverage", "Path Coverage",
        "Coverage Source", "Ghi chú"
    ]
    widths = [16, 38, 42, 16, 10, 10, 22, 22, 24, 24, 20, 20, 18, 48, 64]
    for ci, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 36
    for ci, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = font_white_bold(11)
        cell.fill = header_fill(DARK_BLUE)
        cell.alignment = center()
        cell.border = THIN

    frontend_summary = load_frontend_coverage_summary()
    backend_summary = load_backend_coverage_json()
    for row_idx, func in enumerate(functions, start=2):
        coverage = build_coverage_info(func, frontend_summary, backend_summary)
        passed = sum(1 for r in func["results"] if r["pf"] == "P")
        failed = sum(1 for r in func["results"] if r["pf"] == "F")
        row_data = [
            func["code"],
            func["name"],
            TECHNIQUE_MAP.get(func["code"], ""),
            len(func["utcids"]),
            passed,
            failed,
            format_coverage(coverage["test_design"]),
            format_coverage(coverage["blackbox"]),
            format_coverage(coverage["line"]),
            format_coverage(coverage["branch"]),
            format_coverage(coverage["statement"]),
            format_coverage(coverage["condition"]),
            format_coverage(coverage["path"]),
            coverage["source"],
            coverage["note"],
        ]

        ws.row_dimensions[row_idx].height = 54
        for ci, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=value)
            cell.font = font_normal(10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = THIN
            if ci in (5, 7, 8, 9, 10, 11, 12, 13):
                cell.alignment = center()
            if ci == 5 and failed == 0:
                cell.fill = header_fill("F0FFF0")
            if ci == 6 and failed > 0:
                cell.fill = header_fill("FFF0F0")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ─── MAIN ───────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Sheet1 đầu tiên (active sheet)
    build_sheet1(wb, FUNCTIONS)

    # Coverage summary sheet for black-box/white-box evidence
    build_coverage_sheet(wb, FUNCTIONS)

    # Function sheets
    for i, func in enumerate(FUNCTIONS, start=1):
        sheet_name = f"Function{i}"
        build_function_sheet(wb, func, sheet_name)

    out_path = r"C:\Users\Gnurt\Desktop\KCPM\UnitestKCPM.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        # File is open in Excel, save with suffix
        out_path = r"C:\Users\Gnurt\Desktop\KCPM\UnitestKCPM_new.xlsx"
        wb.save(out_path)
        print(f"[WARNING] File goc dang mo, da luu: {out_path}")

    total_tc = sum(len(f["utcids"]) for f in FUNCTIONS)
    total_pass = sum(sum(1 for r in f["results"] if r["pf"] == "P") for f in FUNCTIONS)
    total_fail = total_tc - total_pass
    print(f"[OK] Generated: {out_path}")
    print(f"   Sheets: Sheet1 + Coverage + {len(FUNCTIONS)} Function sheets")
    print(f"   Total test cases: {total_tc} | Passed: {total_pass} | Failed: {total_fail}")
    print(f"   Members: Nguyen Chi Trung, Minh Phung, Nguyen Hoang Phung, Dang, Thanh Duy")


if __name__ == "__main__":
    main()
